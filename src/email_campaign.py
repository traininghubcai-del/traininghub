"""Email campaign: per-student reminder emails 7 / 3 / 1 days before class.

Pipeline:
  1. ``generate_emails()`` — for every registration with an upcoming class,
     render one branded HTML email per reminder stage (7/3/1 days out, same
     design language as the landing page) plus an ``.ics`` calendar invite,
     into ``data/email_campaign/emails/``. File names start with the
     registration id. The full plan is written to ``campaign_schedule.xlsx``.
  2. ``run_sender(today)`` — finds the stage emails due on ``today`` (or
     overdue and never sent) and sends them through the Brevo SMTP relay,
     appending one row per email to ``outbox.xlsx``. That ledger is what makes
     the run re-runnable: it is read back by ``sent_keys()`` to decide what has
     already gone out. With ``EMAIL_SEND_ENABLED`` off the whole path still
     runs and every row lands as SIMULATED, so a demo costs no mail — and
     burns those reminders, since a simulated row still counts as sent.
     Settings and the two switches: ``docs/RULE-email-sending.md``.

Reads the same sources as the app: the SQLite repo (registrations) and
events.xlsx via src.catalog. Writes only inside data/email_campaign/.
"""
import base64
import json
import os
import smtplib
import ssl
import threading
import time
from datetime import date, datetime, timedelta
from email.message import EmailMessage
from email.utils import formataddr
from urllib.error import HTTPError, URLError
from urllib.parse import quote
from urllib.request import Request, urlopen

from openpyxl import Workbook, load_workbook

from config import (BREVO_API_KEY, BREVO_API_URL, CAMPAIGN_OUTBOX_XLSX,
                    CAMPAIGN_SCHEDULE_XLSX, EMAIL_FROM,
                    EMAIL_FROM_NAME, EMAIL_OUT_DIR, EMAIL_REPLY_TO,
                    EMAIL_SEND_ENABLED, EMAIL_SEND_PAUSE, REMINDER_DAYS,
                    SMTP_HOST, SMTP_PASSWORD, SMTP_PORT, SMTP_SSL, SMTP_USER)
from src.catalog import event_view, is_active, load_catalog
from src.db import get_repository

# --- the three reminder stages -------------------------------------------------
# Same email skeleton, different urgency. `accent` colors the countdown chip.
# Stage 0 is not a reminder — it's the receipt sent the moment someone
# registers. It rides the same skeleton so the confirmation and the reminders
# are visibly one family of mail, and so a template fix lands on all four.
CONFIRM_STAGE = 0

STAGES = {
    CONFIRM_STAGE: {
        "eyebrow": "YOU'RE REGISTERED",
        "chip": "SEATS CONFIRMED",
        "headline": "YOU'RE IN — SEATS CONFIRMED",
        "lead": ("Thanks — your team's seats are reserved. Everything you need is "
                 "below; add it to your calendar now so the date is locked in. "
                 "We'll email you again as the class gets close."),
        "accent": "#1f8a4c",
    },
    7: {
        "eyebrow": "ONE WEEK OUT",
        "chip": "7 DAYS TO GO",
        "headline": "YOUR CLASS IS IN 7 DAYS",
        "lead": ("Your team's seats are confirmed. Lock the date in now — "
                 "one tap below adds the class to your calendar."),
        "accent": "#1d6fc2",
    },
    3: {
        "eyebrow": "GETTING CLOSE",
        "chip": "3 DAYS TO GO",
        "headline": "3 DAYS UNTIL YOUR CLASS",
        "lead": ("Quick reminder — here are your class details and the team "
                 "you registered. Reply to this email if anything changed."),
        "accent": "#e0a92b",
    },
    1: {
        "eyebrow": "TOMORROW",
        "chip": "TOMORROW",
        "headline": "1 DAY — BE PREPARED!",
        "lead": ("Your class is tomorrow. Run through the checklist below so "
                 "your team gets the most out of the day."),
        "accent": "#1f8a4c",
    },
}

SCHEDULE_COLUMNS = ["reg_id", "contact_email", "company_name", "event_id", "topic",
                    "class_date", "stage_days", "send_on", "subject",
                    "html_file", "ics_file"]
OUTBOX_COLUMNS = ["sent_at", "simulated_today", "stage_days", "send_on", "reg_id",
                  "contact_email", "company_name", "event_id", "topic",
                  "class_date", "subject", "html_file", "status"]


def _subject(stage, view):
    if stage == CONFIRM_STAGE:
        return f"You're registered: {view['topic']} — {view['date_display']}"
    if stage == 1:
        return f"Tomorrow: {view['topic']} — be prepared!"
    return f"{stage} days to go: {view['topic']} — {view['date_display']}"


# --- recipients ------------------------------------------------------------------

def upcoming_jobs(today=None):
    """One job per (registration x stage) for classes on/after `today`.

    Joins the repo's flat registrations to the live events.xlsx catalog, so a
    date change in Excel automatically reschedules every reminder.
    """
    today = today or date.today()
    events, _, _ = load_catalog()
    jobs = []
    for reg in get_repository().all_registrations_flat():
        ev = events.get(reg["event_id"])
        if not ev or not is_active(ev):
            continue
        view = event_view(ev)
        class_date = datetime.strptime(str(ev["event_date"]), "%Y-%m-%d").date()
        if class_date < today:
            continue  # class already happened — nothing left to remind
        reg_id = int(reg.get("id") or reg.get("reg_id") or 0)
        for stage in REMINDER_DAYS:
            jobs.append({
                "reg_id": reg_id,
                "reg": reg,
                "event": ev,
                "view": view,
                "class_date": class_date,
                "stage": stage,
                "send_on": class_date - timedelta(days=stage),
                "subject": _subject(stage, view),
                "html_name": f"{reg_id:04d}_{stage}day.html",
                "ics_name": f"{reg_id:04d}_invite.ics",
            })
    return jobs


# --- calendar (the "add to calendar" button) -------------------------------------

def _dt_compact(ev, key):
    """'2026-06-17' + '09:00' -> '20260617T090000' (floating local time)."""
    d = str(ev["event_date"]).replace("-", "")
    t = str(ev[key]).replace(":", "")
    return f"{d}T{t}00"


def google_calendar_url(ev, view):
    details = f"{view['host_label']}\nRegistered via the M&A Supply Training Hub."
    return ("https://calendar.google.com/calendar/render?action=TEMPLATE"
            f"&text={quote(view['topic'] + ' — M&A Supply Training')}"
            f"&dates={_dt_compact(ev, 'start_time')}/{_dt_compact(ev, 'end_time')}"
            f"&location={quote(view['event_location'])}"
            f"&details={quote(details)}")


def ics_text(ev, view, reg_id):
    """Minimal .ics invite (floating local time — fine for a same-state class)."""
    stamp = datetime.now().strftime("%Y%m%dT%H%M%S")
    return "\r\n".join([
        "BEGIN:VCALENDAR",
        "VERSION:2.0",
        "PRODID:-//M&A Supply//Training Hub//EN",
        "BEGIN:VEVENT",
        f"UID:reg-{reg_id}-{ev['event_id']}@ma-supply-training",
        f"DTSTAMP:{stamp}",
        f"DTSTART:{_dt_compact(ev, 'start_time')}",
        f"DTEND:{_dt_compact(ev, 'end_time')}",
        f"SUMMARY:{view['topic']} — M&A Supply Training",
        f"LOCATION:{view['event_location']}",
        f"DESCRIPTION:{view['host_label']}",
        "END:VEVENT",
        "END:VCALENDAR",
        "",
    ])


# --- HTML rendering (same design language as static/styles.css) -------------------
# Email clients ignore <style> blocks and custom fonts, so this is the site's
# palette re-built as a 600px table with inline styles + system font stacks.

_FONT = "'Saira','Barlow',Arial,Helvetica,sans-serif"


def _btn(href, label, solid=True):
    style = (f"background:#1d6fc2;color:#ffffff;border:2px solid #1d6fc2;" if solid else
             f"background:#ffffff;color:#155a9f;border:2px solid #1d6fc2;")
    return (f'<a href="{href}" style="display:inline-block;{style}'
            f'font-family:{_FONT};font-weight:700;font-size:14px;letter-spacing:.02em;'
            f'text-decoration:none;padding:13px 26px;margin:4px 6px;">{label}</a>')


def _detail_row(label, value):
    return (f'<tr><td style="padding:7px 0;font-family:{_FONT};font-weight:600;'
            f'font-size:12px;letter-spacing:.06em;color:#0a2540;text-transform:uppercase;'
            f'width:120px;vertical-align:top;">{label}</td>'
            f'<td style="padding:7px 0;font-family:Barlow,Arial,sans-serif;font-size:15px;'
            f'color:#16222f;">{value}</td></tr>')


def _prep_checklist(view):
    items = [
        "Arrive 15 minutes early — check-in opens before the first session.",
        "Bring a notepad; bring your tools/gauges if your class is hands-on.",
        "Make sure every attendee on your list below knows the time and location.",
    ]
    if view.get("notes"):
        items.insert(0, f"<b>From your branch:</b> {view['notes']}")
    lis = "".join(
        f'<li style="margin:6px 0;color:#6b5212;font-family:Barlow,Arial,sans-serif;'
        f'font-size:14px;">{i}</li>' for i in items)
    return (f'<div style="background:#fff8e6;border-left:4px solid #e0a92b;'
            f'padding:14px 18px;margin:22px 0 4px;">'
            f'<p style="margin:0 0 6px;font-family:{_FONT};font-weight:700;font-size:13px;'
            f'letter-spacing:.08em;color:#6b5212;text-transform:uppercase;">Be prepared</p>'
            f'<ul style="margin:0;padding-left:18px;">{lis}</ul></div>')


def _footer_note(stage):
    if stage == CONFIRM_STAGE:
        return "Registration confirmation. Reply to this email to change your team."
    return (f"Automated class reminder ({stage} day"
            f"{'s' if stage != 1 else ''} before your class).")


def render_email(job):
    """One full HTML email for a (registration x stage) job."""
    reg, view, ev, stage = job["reg"], job["view"], job["event"], job["stage"]
    s = STAGES[stage]
    cal_url = google_calendar_url(ev, view)
    note = (f'<tr><td style="padding:0 28px 6px;">'
            f'<p style="margin:0;padding:10px 14px;background:#fff8e6;'
            f'border-left:4px solid #e0a92b;color:#6b5212;font-family:Barlow,Arial,'
            f'sans-serif;font-size:13px;">{view["notes"]}</p></td></tr>'
            if view.get("notes") and stage != 1 else "")
    prep = (f'<tr><td style="padding:0 28px;">{_prep_checklist(view)}</td></tr>'
            if stage == 1 else "")
    attendees = reg.get("attendees") or "(no attendee names on file)"

    return f"""<!doctype html>
<html lang="en"><head><meta charset="utf-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>{job['subject']}</title></head>
<body style="margin:0;padding:0;background:#f3f6fa;">
<table role="presentation" width="100%" cellpadding="0" cellspacing="0" style="background:#f3f6fa;">
<tr><td align="center" style="padding:26px 12px;">
<table role="presentation" width="600" cellpadding="0" cellspacing="0"
       style="width:600px;max-width:100%;background:#ffffff;box-shadow:0 10px 30px rgba(10,37,64,.10);">

  <!-- header : navy bar, blue rule (matches .ma-header) -->
  <tr><td style="background:#0a2540;border-bottom:3px solid #1d6fc2;padding:16px 28px;">
    <span style="font-family:{_FONT};font-weight:800;font-size:17px;letter-spacing:.04em;color:#ffffff;">
      M&amp;A SUPPLY CO. <span style="font-weight:500;color:#acc6e3;">— TRAINING</span>
    </span>
  </td></tr>

  <!-- hero : deep navy, eyebrow + countdown (matches .ma-hero) -->
  <tr><td style="background:#06192e;padding:34px 28px 30px;">
    <p style="margin:0 0 12px;font-family:{_FONT};font-weight:700;font-size:11px;
              letter-spacing:.22em;color:#6fb0ee;">{s['eyebrow']}</p>
    <h1 style="margin:0 0 14px;font-family:{_FONT};font-weight:800;font-size:30px;
               line-height:1.05;color:#ffffff;text-transform:uppercase;">{s['headline']}</h1>
    <span style="display:inline-block;background:{s['accent']};color:#ffffff;
                 font-family:{_FONT};font-weight:800;font-size:13px;letter-spacing:.08em;
                 padding:7px 16px;">{s['chip']} &middot; {view['weekday_display'].upper()}, {view['date_display'].upper()}</span>
    <p style="margin:16px 0 0;font-family:Barlow,Arial,sans-serif;font-size:15px;
              line-height:1.55;color:#c7d8e8;">{s['lead']}</p>
  </td></tr>

  <!-- class card : blue left rail (matches .ma-class-banner) -->
  <tr><td style="padding:26px 28px 8px;">
    <table role="presentation" width="100%" cellpadding="0" cellspacing="0"
           style="border:1px solid #d9e2ec;border-left:5px solid #1d6fc2;">
      <tr><td style="padding:18px 20px;">
        <h2 style="margin:0 0 12px;font-family:{_FONT};font-weight:800;font-size:20px;
                   color:#0a2540;text-transform:uppercase;line-height:1.1;">{view['topic']}</h2>
        <table role="presentation" width="100%" cellpadding="0" cellspacing="0">
          {_detail_row("Date", f"{view['weekday_display']}, {view['date_display']}")}
          {_detail_row("Time", view['time_display'])}
          {_detail_row("Location", view['event_location'])}
          {_detail_row("Hosted by", view['host_label'])}
        </table>
      </td></tr>
    </table>
  </td></tr>

  {note}{prep}

  <!-- add to calendar -->
  <tr><td align="center" style="padding:18px 28px 6px;">
    {_btn(cal_url, "&#128197;&nbsp; ADD TO GOOGLE CALENDAR")}
    {_btn("cid:invite.ics", "DOWNLOAD .ICS (OUTLOOK / APPLE)", solid=False)}
    <p style="margin:10px 0 0;font-family:Barlow,Arial,sans-serif;font-size:12px;color:#56697a;">
      The attached <b>invite.ics</b> works with Outlook and Apple Calendar.</p>
  </td></tr>

  <!-- registered team (matches .ma-tag / blue-soft) -->
  <tr><td style="padding:16px 28px 26px;">
    <table role="presentation" width="100%" cellpadding="0" cellspacing="0"
           style="background:#e9f1fa;border:1px solid #cfe1f5;">
      <tr><td style="padding:16px 20px;">
        <p style="margin:0 0 8px;font-family:{_FONT};font-weight:700;font-size:12px;
                  letter-spacing:.08em;color:#155a9f;text-transform:uppercase;">
          Your registered team — {reg['company_name']}</p>
        <p style="margin:0;font-family:Barlow,Arial,sans-serif;font-size:14px;color:#16222f;">
          {attendees}</p>
        <p style="margin:8px 0 0;font-family:Barlow,Arial,sans-serif;font-size:12px;color:#56697a;">
          {reg['num_attending']} attending &middot; registered with {reg['contact_email']}
          &middot; reply to this email to make changes.</p>
      </td></tr>
    </table>
  </td></tr>

  <!-- footer (matches .ma-footer) -->
  <tr><td style="background:#06192e;padding:18px 28px;">
    <p style="margin:0;font-family:Barlow,Arial,sans-serif;font-size:12px;color:#b9cadb;">
      M&amp;A Supply Company — Training Hub &middot; {_footer_note(stage)}</p>
  </td></tr>

</table>
</td></tr></table>
</body></html>"""


# --- the registration receipt ------------------------------------------------------

def confirmation_job(reg_id, reg, attendees):
    """Build the stage-0 job for a registration that was just saved.

    Takes the values the request already has in hand rather than re-reading the
    table: the confirmation is sent on the write path, so a full scan per signup
    would get more expensive with every registration ever taken.

    Returns None when the class can't be resolved — an unknown or deactivated
    event is not worth failing a completed registration over.
    """
    events, _, _ = load_catalog()
    ev = events.get(reg["event_id"])
    if not ev or not is_active(ev):
        return None
    view = event_view(ev)
    joined = ", ".join(f"{a['name']} ({a['role']})" if a.get("role") else a["name"]
                       for a in attendees) or "(no attendee names on file)"
    return {
        "reg_id": reg_id,
        # the flat shape the templates expect, same keys all_registrations_flat gives
        "reg": {"contact_email": reg["contact_email"],
                "company_name": reg["company_name"],
                "num_attending": reg["num_attending"],
                "attendees": joined},
        "event": ev,
        "view": view,
        "class_date": datetime.strptime(str(ev["event_date"]), "%Y-%m-%d").date(),
        "stage": CONFIRM_STAGE,
        "send_on": date.today(),
        "subject": _subject(CONFIRM_STAGE, view),
        "html_name": f"{reg_id:04d}_confirmation.html",
        "ics_name": f"{reg_id:04d}_invite.ics",
    }


def send_confirmation(reg_id, reg, attendees):
    """Send the "you're registered" receipt. Never raises.

    The seats are already committed by the time this runs, so nothing here is
    allowed to turn a successful registration into an error the dealer sees.
    Delivery goes through send_jobs() so the receipt lands in the same outbox
    ledger as the reminders and is auditable next to them.
    """
    try:
        job = confirmation_job(reg_id, reg, attendees)
        if job is None:
            return "SKIPPED — class not found or inactive"
        # keep the copy the dealer was sent, so the outbox's html_file column
        # points at something real for receipts the way it does for reminders
        try:
            EMAIL_OUT_DIR.mkdir(parents=True, exist_ok=True)
            (EMAIL_OUT_DIR / job["html_name"]).write_text(render_email(job), encoding="utf-8")
        except OSError:
            pass          # a read-only disk must not cost the dealer their receipt
        send_jobs([job], date.today())
        return job.get("status", "UNKNOWN")
    except Exception as e:  # noqa: BLE001 - a mail fault must not surface as a failed signup
        return f"FAILED — {type(e).__name__}: {e}"


# --- step 1: generate ---------------------------------------------------------------

def generate_emails(today=None):
    """Render every (registration x stage) email + one .ics per registration.

    Returns the job list. Also rewrites campaign_schedule.xlsx — the full plan,
    one row per email, so the campaign can be reviewed in Excel before any send.
    """
    EMAIL_OUT_DIR.mkdir(parents=True, exist_ok=True)
    jobs = upcoming_jobs(today)
    ics_done = set()
    for job in jobs:
        (EMAIL_OUT_DIR / job["html_name"]).write_text(render_email(job), encoding="utf-8")
        if job["reg_id"] not in ics_done:
            (EMAIL_OUT_DIR / job["ics_name"]).write_text(
                ics_text(job["event"], job["view"], job["reg_id"]), encoding="utf-8")
            ics_done.add(job["reg_id"])

    wb = Workbook()
    ws = wb.active
    ws.title = "schedule"
    ws.append(SCHEDULE_COLUMNS)
    for j in sorted(jobs, key=lambda j: (j["send_on"], j["reg_id"], -j["stage"])):
        ws.append([j["reg_id"], j["reg"]["contact_email"], j["reg"]["company_name"],
                   j["view"]["event_id"], j["view"]["topic"], str(j["class_date"]),
                   j["stage"], str(j["send_on"]), j["subject"],
                   f"emails/{j['html_name']}", f"emails/{j['ics_name']}"])
    CAMPAIGN_SCHEDULE_XLSX.parent.mkdir(parents=True, exist_ok=True)
    wb.save(CAMPAIGN_SCHEDULE_XLSX)
    return jobs


# --- step 2: send -------------------------------------------------------------------

def _plain_text(job):
    """Text fallback for clients that won't render HTML. Every mail is sent
    multipart/alternative — a text-only reader still gets the what/when/where."""
    v = job["view"]
    return (f"{job['subject']}\n\n"
            f"{v['topic']}\n"
            f"{v.get('weekday', '')} {v.get('event_date', '')}  {v.get('time_display', '')}\n"
            f"{v.get('address_display') or v.get('event_location', '')}\n\n"
            f"Registered under: {job['reg']['company_name']}\n\n"
            "Can't make it? Let your M&A branch know at least 24 hours ahead so "
            "the seat can go to another dealer.\n")


def transport():
    """Which way mail goes out: "api" when a Brevo API key is set, else "smtp".

    Not a preference — a deployment fact. Railway blocks outbound SMTP on every
    port, so on the live host the API is the only transport that works at all.
    SMTP stays supported because it is what runs locally and what a different
    host would use.
    """
    return "api" if BREVO_API_KEY else "smtp"


def missing_settings():
    """Which mail settings are unset, in the words the person setting them uses.

    One list, so the send path, the /api/mail-status endpoint and tools/mail_test.py
    all say the same thing about the same misconfiguration instead of each
    inventing its own phrasing.
    """
    missing = []
    if transport() == "api":
        if not EMAIL_FROM:
            missing.append("EMAIL_FROM")
        return missing
    if not SMTP_HOST:
        missing.append("SMTP_HOST")
    if not SMTP_USER:
        missing.append("SMTP_USER (the Brevo SMTP login, not the from address)")
    if not SMTP_PASSWORD:
        missing.append("SMTP_PASSWORD (the Brevo SMTP key, xsmtpsib-...)")
    if not EMAIL_FROM:
        missing.append("EMAIL_FROM")
    return missing


def api_send(job, to):
    """POST one email to Brevo's transactional endpoint. Raises on rejection.

    Same three parts as the SMTP path — text, HTML, the .ics — so a dealer
    cannot tell which transport carried their mail.
    """
    ics = ics_text(job["event"], job["view"], job["reg_id"]).encode("utf-8")
    payload = {
        "sender": {"name": EMAIL_FROM_NAME, "email": EMAIL_FROM},
        "to": [{"email": to}],
        "replyTo": {"email": EMAIL_REPLY_TO},
        "subject": job["subject"],
        "textContent": _plain_text(job),
        "htmlContent": render_email(job),
        "attachment": [{"content": base64.b64encode(ics).decode("ascii"),
                        "name": job["ics_name"]}],
    }
    req = Request(BREVO_API_URL, method="POST",
                  data=json.dumps(payload).encode("utf-8"),
                  headers={"api-key": BREVO_API_KEY, "content-type": "application/json",
                           "accept": "application/json"})
    with urlopen(req, timeout=30) as r:
        return json.loads(r.read() or b"{}")


def api_probe():
    """Check the API key without sending. Returns a status string, never raises.

    GET /v3/account is the cheapest authenticated call Brevo has, so this costs
    no email credit and puts nothing in front of a dealer.
    """
    if not BREVO_API_KEY:
        return "BREVO_API_KEY is not set"
    req = Request("https://api.brevo.com/v3/account",
                  headers={"api-key": BREVO_API_KEY, "accept": "application/json"})
    try:
        with urlopen(req, timeout=20) as r:
            acct = json.loads(r.read() or b"{}")
        who = acct.get("email") or acct.get("companyName") or "authenticated"
        return f"OK — Brevo API key valid ({who})"
    except HTTPError as e:
        body = e.read().decode("utf-8", "replace")[:200]
        if e.code == 401:
            return ("FAILED — Brevo rejected the API key (401). BREVO_API_KEY must be "
                    "an API key from Brevo -> SMTP & API -> API keys, starting "
                    f"xkeysib-, not the SMTP key. {body}")
        return f"FAILED — Brevo API returned {e.code}: {body}"
    except URLError as e:
        return f"FAILED — could not reach the Brevo API: {e.reason}"


def smtp_connect():
    """One authenticated connection to the relay. Raises on failure.

    Separated from sending so the diagnostic tool can prove the credentials
    work without putting a message in front of a dealer.
    """
    ctx = ssl.create_default_context()
    if SMTP_SSL:
        conn = smtplib.SMTP_SSL(SMTP_HOST, SMTP_PORT, timeout=30, context=ctx)
    else:
        conn = smtplib.SMTP(SMTP_HOST, SMTP_PORT, timeout=30)
        conn.ehlo()
        conn.starttls(context=ctx)
        conn.ehlo()
    conn.login(SMTP_USER, SMTP_PASSWORD)
    return conn


def _auth_hint(e):
    """Turn a 535 into the thing that is actually wrong.

    Brevo returns the same "authentication failed" for a bad key and for the
    single most common mistake — putting the sending address in SMTP_USER when
    Brevo issues its own login. Guessing wrong here costs an afternoon, so the
    status string names both and says where to look.
    """
    code = getattr(e, "smtp_code", "")
    detail = getattr(e, "smtp_error", b"")
    detail = detail.decode("utf-8", "replace") if isinstance(detail, bytes) else str(detail)
    hint = ("check SMTP_USER is the Brevo SMTP login from Brevo -> SMTP & API "
            "(often 9a1b2c001@smtp-brevo.com), NOT the from address, and that "
            "SMTP_PASSWORD is the SMTP key")
    return f"FAILED — {SMTP_HOST} rejected the login ({code} {detail.strip()}); {hint}"


def _build(job, to):
    """The multipart message for one job: text, HTML, and the .ics invite."""
    msg = EmailMessage()
    msg["Subject"] = job["subject"]
    msg["From"] = formataddr((EMAIL_FROM_NAME, EMAIL_FROM))
    msg["To"] = to
    msg["Reply-To"] = EMAIL_REPLY_TO
    msg.set_content(_plain_text(job))
    msg.add_alternative(render_email(job), subtype="html")
    msg.add_attachment(
        ics_text(job["event"], job["view"], job["reg_id"]).encode("utf-8"),
        maintype="text", subtype="calendar", filename=job["ics_name"])
    return msg


def _deliver(job, conn=None):
    """Send one email through the SMTP relay, with the .ics attached.

    Returns the string that lands in the outbox `status` column. That string is
    load-bearing: sent_keys() treats a row as "already sent" unless the status
    says it failed, so a bounce or an auth error retries on the next run instead
    of being silently swallowed. Which is also why every failure path here says
    what went wrong in full — the outbox is the only place that text survives.

    Nothing leaves the machine unless EMAIL_SEND_ENABLED is on AND the settings
    are complete — either missing and this stays a simulation.

    `conn` is a live connection from send_jobs(), reused across a batch. Without
    one this opens and closes its own, so a single receipt still sends.
    """
    if not EMAIL_SEND_ENABLED:
        return "SIMULATED — written to outbox, no email left this machine"
    missing = missing_settings()
    if missing:
        return f"FAILED — not configured: {', '.join(missing)}; nothing was sent"

    to = str(job["reg"].get("contact_email", "")).strip()
    if not to:
        return "FAILED — registration has no contact email"

    try:
        if transport() == "api":
            api_send(job, to)
        elif conn is not None:
            conn.send_message(_build(job, to))
        else:
            with smtp_connect() as s:
                s.send_message(_build(job, to))
        if EMAIL_SEND_PAUSE > 0:
            time.sleep(EMAIL_SEND_PAUSE)     # stay under the relay's burst limits
        return f"SENT to {to}"
    except HTTPError as e:
        # Brevo says why in the body, and that sentence is the whole diagnosis:
        # a bad key, an unverified sender and a malformed address all arrive
        # here as a 400-class code that means nothing on its own.
        body = e.read().decode("utf-8", "replace")[:300]
        return f"FAILED — Brevo API {e.code}: {body}"
    except URLError as e:
        return f"FAILED — could not reach the Brevo API: {e.reason}"
    except smtplib.SMTPAuthenticationError as e:
        return _auth_hint(e)
    except smtplib.SMTPSenderRefused as e:
        # Login worked, the From address did not. On Brevo that is a sender that
        # was never verified — the message is refused after a clean handshake,
        # which looks nothing like an auth problem from the outside.
        return (f"FAILED — {EMAIL_FROM} was refused as the sender ({e.smtp_code} "
                f"{e.smtp_error!r}); verify it under Brevo -> Senders")
    except smtplib.SMTPRecipientsRefused as e:
        return f"FAILED — {to} was refused: {e.recipients}"
    except Exception as e:  # noqa: BLE001 - one bad address must not stop the batch
        return f"FAILED — {type(e).__name__}: {e}"


def _row_stage(rec):
    """The stage on one outbox row, or None when it can't be read.

    outbox.xlsx is a spreadsheet a person can open and edit, so a blank or
    re-typed cell must not raise in the middle of a send.
    """
    try:
        return int(rec.get("stage_days"))
    except (TypeError, ValueError):
        return None


def sent_keys():
    """(reg_id, stage) pairs already sent, from outbox.xlsx — the dedupe ledger.

    A row whose status starts with FAILED does NOT count: that email never
    reached anyone, so the next run should try it again. Anything else (a real
    SENT, or a SIMULATED row from demo mode) counts as done and is never
    re-sent — that is what stops a dealer getting the same reminder twice.
    """
    if not CAMPAIGN_OUTBOX_XLSX.exists():
        return set()
    ws = load_workbook(CAMPAIGN_OUTBOX_XLSX, read_only=True).active
    keys = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        rec = dict(zip(OUTBOX_COLUMNS, row))
        if rec.get("reg_id") is None:
            continue
        if str(rec.get("status") or "").strip().upper().startswith("FAILED"):
            continue
        keys.add((int(rec["reg_id"]), int(rec["stage_days"])))
    return keys


def reminded_today(event_id, today=None):
    """reg_ids that already got a REMINDER for this class today.

    The (reg, stage) ledger alone can't stop a double-click on the Admin
    "send now" button: the second press skips the stage it just sent and falls
    through to the next one, which is a different key but the same dealer
    getting mail twice in a row. Same-day is the guard that matches what a
    person would call a duplicate.

    Registration receipts share this ledger but are NOT reminders: someone who
    signed up this morning has had no reminder at all, and counting their
    receipt here would silently drop them from today's send.
    """
    today = today or date.today()
    if not CAMPAIGN_OUTBOX_XLSX.exists():
        return set()
    ws = load_workbook(CAMPAIGN_OUTBOX_XLSX, read_only=True).active
    out = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        rec = dict(zip(OUTBOX_COLUMNS, row))
        if str(rec.get("event_id") or "") != str(event_id):
            continue
        if _row_stage(rec) == CONFIRM_STAGE:
            continue
        if str(rec.get("status") or "").strip().upper().startswith("FAILED"):
            continue
        if str(rec.get("sent_at") or "")[:10] == str(today):
            out.add(int(rec["reg_id"]))
    return out


def due_jobs(today=None):
    """Emails whose send day arrived (or passed) and were never sent,
    for classes still in the future."""
    today = today or date.today()
    done = sent_keys()
    return [j for j in upcoming_jobs(today)
            if j["send_on"] <= today and (j["reg_id"], j["stage"]) not in done]


def run_sender(today=None):
    """'Send' everything due as of `today`: one outbox.xlsx row per email.

    Re-runnable any day — already-sent (reg, stage) pairs are skipped, so this
    is exactly the function a daily cron/scheduler will call after deploy.
    """
    today = today or date.today()
    return send_jobs(due_jobs(today), today)


# The outbox is a single .xlsx rewritten whole on every append. Registration
# receipts are sent from request threads, so two dealers signing up at once
# would otherwise read-modify-write the same workbook concurrently: rows vanish,
# and a half-finished save leaves a file that sent_keys() can no longer load —
# taking the reminder dedupe down with it.
_OUTBOX_LOCK = threading.Lock()


def send_jobs(jobs, today=None):
    """Deliver these jobs and append one outbox row each.

    The daily scheduler, the Admin lens "send now" button and the registration
    receipt all come through here, so they share one ledger: a reminder sent by
    hand this afternoon is already recorded when tonight's run computes what's
    due, and the dealer never gets it twice.
    """
    today = today or date.today()
    if not jobs:
        return []

    # Deliver first, outside the lock: SMTP is slow and one dealer's send must
    # not hold the ledger against everyone else's.
    #
    # One connection for the whole batch. A day's reminders reconnecting and
    # re-authenticating per message is what trips a relay's connection limits
    # first — well before the message count matters. If the connection can't be
    # opened at all, every job still goes through _deliver() to get the real
    # reason written into its own outbox row, rather than the batch dying with
    # one error nobody can see.
    now = datetime.now().isoformat(timespec="seconds")
    conn = None
    if EMAIL_SEND_ENABLED and transport() == "smtp" and not missing_settings():
        try:
            conn = smtp_connect()
        except Exception:  # noqa: BLE001 - _deliver reports it per job
            conn = None
    try:
        for j in jobs:
            j["status"] = _deliver(j, conn)
            # A relay that hangs up mid-batch — idle timeout, a rate limit, a
            # restart on their side — would otherwise fail every remaining job
            # on a connection that is already gone. Reconnect once and retry
            # this one; if that fails too, its FAILED row retries tomorrow.
            if conn is not None and "SMTPServerDisconnected" in str(j["status"]):
                try:
                    conn = smtp_connect()
                except Exception:  # noqa: BLE001 - keep the original failure text
                    conn = None
                else:
                    j["status"] = _deliver(j, conn)
    finally:
        if conn is not None:
            try:
                conn.quit()
            except Exception:  # noqa: BLE001 - the mail is already sent
                pass

    with _OUTBOX_LOCK:
        if CAMPAIGN_OUTBOX_XLSX.exists():
            wb = load_workbook(CAMPAIGN_OUTBOX_XLSX)
            ws = wb.active
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = "outbox"
            ws.append(OUTBOX_COLUMNS)
        for j in jobs:
            ws.append([now, str(today), j["stage"], str(j["send_on"]), j["reg_id"],
                       j["reg"]["contact_email"], j["reg"]["company_name"],
                       j["view"]["event_id"], j["view"]["topic"], str(j["class_date"]),
                       j["subject"], f"emails/{j['html_name']}", j["status"]])
        CAMPAIGN_OUTBOX_XLSX.parent.mkdir(parents=True, exist_ok=True)
        wb.save(CAMPAIGN_OUTBOX_XLSX)
    return jobs


# --- diagnostics -------------------------------------------------------------

def mail_status(probe=False):
    """Why mail is or isn't leaving this server.

    Booleans for anything secret: this is a read-only health check, never a way
    to read a credential back out. The outbox tail is the useful part — it
    carries the provider's own rejection text, which is invisible from the UI.

    `probe` opens a real connection and logs in, then hangs up without sending.
    That is the one question the static fields can't answer — whether the login
    and the key actually match — and the answer is otherwise only reachable by
    mailing a live dealer and reading the outbox afterwards.
    """
    from config import DATA
    raw = os.environ.get("EMAIL_SEND_ENABLED", "")
    info = {
        "send_enabled": EMAIL_SEND_ENABLED,
        "send_enabled_raw": repr(raw),          # catches "True "/"yes please"/typos
        # Which of the two paths below is actually carrying mail. Everything
        # after it describes the one that isn't, half the time.
        "transport": transport(),
        "brevo_api_key_set": bool(BREVO_API_KEY),
        "smtp_host": SMTP_HOST,
        "smtp_port": SMTP_PORT,
        # The login is shown in full, unlike the key. It is a username, not a
        # secret, and it is the field that is actually wrong most of the time:
        # a boolean here would hide a Gmail address sitting where Brevo's own
        # login belongs, which is the exact failure this endpoint exists for.
        "smtp_user": SMTP_USER,
        "smtp_user_set": bool(SMTP_USER),
        "smtp_password_set": bool(SMTP_PASSWORD),
        "smtp_password_looks_like_brevo_key": SMTP_PASSWORD.startswith("xsmtpsib-"),
        "missing": missing_settings(),
        "email_from": EMAIL_FROM,               # must be a validated sender
        "reply_to": EMAIL_REPLY_TO,
        "data_dir": str(DATA),
        "data_dir_exists": DATA.exists(),
        "outbox_exists": CAMPAIGN_OUTBOX_XLSX.exists(),
        "outbox_rows": 0,
        "recent": [],
    }
    try:
        info["data_dir_writable"] = os.access(DATA, os.W_OK)
    except OSError:
        info["data_dir_writable"] = False

    if probe:
        if missing_settings():
            info["login_probe"] = f"not configured: {', '.join(missing_settings())}"
        elif transport() == "api":
            info["login_probe"] = api_probe()
        else:
            try:
                smtp_connect().quit()
                info["login_probe"] = f"OK — logged in to {SMTP_HOST}:{SMTP_PORT}"
            except smtplib.SMTPAuthenticationError as e:
                info["login_probe"] = _auth_hint(e)
            except Exception as e:  # noqa: BLE001 - a health check never raises
                info["login_probe"] = f"FAILED — {type(e).__name__}: {e}"
    if CAMPAIGN_OUTBOX_XLSX.exists():
        try:
            ws = load_workbook(CAMPAIGN_OUTBOX_XLSX, read_only=True).active
            rows = [r for r in ws.iter_rows(min_row=2, values_only=True) if r and r[0]]
            info["outbox_rows"] = len(rows)
            for r in rows[-8:]:
                rec = dict(zip(OUTBOX_COLUMNS, r))
                info["recent"].append({"sent_at": str(rec.get("sent_at")),
                                       "stage": rec.get("stage_days"),
                                       "status": str(rec.get("status"))})
        except Exception as e:  # noqa: BLE001
            info["outbox_error"] = f"{type(e).__name__}: {e}"
    return info
