"""Reminder history per class — read from the email campaign's own ledger.

Deliberately NOT a new table. src/email_campaign.py already appends one row per
reminder to data/email_campaign/outbox.xlsx (sent_at, event_id, reg_id,
stage_days, status), and that file is the dedupe ledger the sender itself trusts
(`sent_keys()`). So the admin table reads the same rows the sender writes.

That is the whole point: when `_deliver()` stops being a stub and starts talking
to a real provider, it keeps appending to the same outbox — and the "total
reminders sent" / "last reminder" columns start showing real deliveries with no
change here. One ledger, no drift between what was sent and what we report.

Everything degrades quietly: no campaign has ever run -> no file -> zeros.
"""
from datetime import datetime

from config import CAMPAIGN_OUTBOX_XLSX

# outbox column order, mirrored from email_campaign.OUTBOX_COLUMNS
_COLS = ["sent_at", "simulated_today", "stage_days", "send_on", "reg_id",
         "contact_email", "company_name", "event_id", "topic",
         "class_date", "subject", "html_file", "status"]


def _pretty(stamp):
    """'2026-06-06T19:20:40' -> 'Jun 6, 2026'. Falls back to the raw string."""
    s = str(stamp or "").strip()
    if not s:
        return ""
    for fmt in ("%Y-%m-%dT%H:%M:%S", "%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M", "%Y-%m-%d"):
        try:
            return datetime.strptime(s[:len(fmt) + 2].strip(), fmt).strftime("%b %-d, %Y")
        except ValueError:
            continue
    return s[:10]


def reminder_stats():
    """{event_id: {sent, last_sent_at, last_sent_display, stages}} for every class
    that has ever had a reminder go out. One workbook read, not one per class."""
    out = {}
    if not CAMPAIGN_OUTBOX_XLSX.exists():
        return out
    try:
        from openpyxl import load_workbook
        wb = load_workbook(CAMPAIGN_OUTBOX_XLSX, read_only=True, data_only=True)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):
            rec = dict(zip(_COLS, row))
            eid = str(rec.get("event_id") or "").strip()
            if not eid:
                continue
            sent_at = str(rec.get("sent_at") or "").strip()
            slot = out.setdefault(eid, {"runs": set(), "letters": 0,
                                        "last_sent_at": "", "stages": set()})
            # One reminder = one DAY the class was reminded, however many
            # students it covered and however many times staff hit the button.
            # Re-generating today must not inflate the number: it is "we
            # reminded this class today", not "we clicked N times".
            slot["runs"].add(sent_at[:10])
            slot["letters"] += 1
            if sent_at > slot["last_sent_at"]:      # ISO stamps sort lexically
                slot["last_sent_at"] = sent_at
            stage = rec.get("stage_days")
            if stage not in (None, ""):
                slot["stages"].add(int(stage))
        wb.close()
    except Exception:  # noqa: BLE001 - a broken ledger must not break the hub
        return out

    for slot in out.values():
        slot["sent"] = len(slot.pop("runs"))
        slot["last_sent_display"] = _pretty(slot["last_sent_at"])
        slot["stages"] = sorted(slot.pop("stages"), reverse=True)
    return out


def for_class(event_id, stats=None):
    """Reminder history for one class, in the shape the admin table renders."""
    stats = reminder_stats() if stats is None else stats
    s = stats.get(str(event_id).strip())
    if not s:
        return {"sent": 0, "letters": 0, "last_sent_at": "", "last_sent_display": "", "stages": []}
    return dict(s)
