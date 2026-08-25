"""The class catalog, read live from data/events.xlsx.

events.xlsx stays the editable source of truth (John edits it in Excel). Nothing
here writes to it — we only read and derive display fields. The DB keeps a cache
of events for referential integrity (see src/db), but this module is the catalog.
"""
import re
from datetime import datetime

from openpyxl import load_workbook

from config import EVENTS_XLSX, MONTHS


def _rows(ws):
    headers = [c.value for c in ws[1]]
    for row in ws.iter_rows(min_row=2, values_only=True):
        if all(v is None for v in row):
            continue
        yield {h: ("" if v is None else v) for h, v in zip(headers, row)}


def load_catalog():
    """Return (events_by_id, branch_list, branch_to_tm) read fresh from events.xlsx."""
    wb = load_workbook(EVENTS_XLSX, read_only=True, data_only=True)
    events = {}
    for e in _rows(wb["events"]):
        eid = str(e.get("event_id", "")).strip()
        if eid:
            events[eid] = {k: (v if isinstance(v, bool) else str(v).strip()) for k, v in e.items()}
    branch_to_tm = {}
    for b in _rows(wb["branches"]):
        name = str(b.get("branch", "")).strip()
        if name:
            branch_to_tm[name] = str(b.get("territory_manager") or "").strip()
    branches = sorted(branch_to_tm.keys(),
                      key=lambda b: int(re.match(r"\d+", b).group()) if re.match(r"\d+", b) else 0)
    wb.close()
    return events, branches, branch_to_tm


def _parse_date(iso):
    """Strict: raises on anything unreadable. Callers that must not die use
    safe_date() instead."""
    return datetime.strptime(str(iso).strip()[:10], "%Y-%m-%d").date()


def safe_date(value):
    """date, or None. openpyxl hands back '2026-09-01 00:00:00' for a real date
    cell — that is a good date, not a broken one, so it is normalised, not
    rejected. Only a genuinely unreadable value returns None."""
    try:
        return _parse_date(value)
    except (ValueError, TypeError):
        return None


def iso_date(value):
    """'YYYY-MM-DD' for anything readable, else the raw text. This is what every
    comparison against `today` must use — comparing a raw
    '2026-09-01 00:00:00' string works by luck, and a raw '9/1/26' does not."""
    d = safe_date(value)
    return d.isoformat() if d else str(value or "").strip()


def to_compact(t):
    """'09:00' -> '9AM', '15:00' -> '3PM', '09:30' -> '9:30AM'."""
    m = re.match(r"^(\d{1,2}):(\d{2})$", str(t).strip())
    if not m:
        return str(t)
    h, mins = int(m.group(1)), m.group(2)
    ap = "AM" if h < 12 else "PM"
    h12 = h % 12 or 12
    return f"{h12}{ap}" if mins == "00" else f"{h12}:{mins}{ap}"


def to_12h(t):
    m = re.match(r"^(\d{1,2}):(\d{2})$", str(t).strip())
    if not m:
        return str(t)
    h, mins = int(m.group(1)), m.group(2)
    ap = "AM" if h < 12 else "PM"
    return f"{h % 12 or 12}:{mins} {ap}"


def time_display(ev):
    """'9:00 AM – 1:00 PM CST' — the BRANCH's local time, always.

    These are in-person classes at a physical branch. A Little Rock class at
    9:00 Central happens at 9:00 Central for everyone, so the label is fixed and
    is never converted to the viewer's device timezone — a Florida dealer must
    not be shown "10:00 AM" for a Little Rock class and drive up an hour late.
    Same convention airlines use: departure airport's local time.
    """
    start = str(ev.get("start_time", "")).strip()
    end = str(ev.get("end_time", "")).strip()
    if not (start and end):
        return "Time TBD"
    tz = str(ev.get("timezone", "")).strip()
    return f"{to_12h(start)} – {to_12h(end)}" + (f" {tz}" if tz else "")


def _weekday_upper(ev):
    """Prefer the sheet's weekday column; fall back to deriving from the date."""
    wk = str(ev.get("weekday") or "").strip()
    if wk:
        return wk.upper()
    d = safe_date(ev.get("event_date"))
    return d.strftime("%A").upper() if d else ""


def build_date_info(ev):
    """Combined string EXACTLY in John's format:
    'WEDNESDAY, APRIL 29, 2026_9AM- 3PM_ FIT INSTALL & COMMISSIONING _ @ NASHVILLE BRANCH'
    """
    d = safe_date(ev.get("event_date"))
    if not d:
        return f"DATE UNREADABLE_ {ev.get('topic', '')} _ @ {ev.get('event_location', '')}"
    date_part = f"{_weekday_upper(ev)}, {MONTHS[d.month - 1].upper()} {d.day}, {d.year}"
    time_part = f"{to_compact(ev['start_time'])}- {to_compact(ev['end_time'])}"
    return f"{date_part}_{time_part}_ {ev['topic']} _ @ {ev['event_location']}"


def missing_info(ev):
    """Which required-for-publishing fields are still blank/TBD, derived live so it
    stays accurate after manual /manage edits. Returns a list like ['time','trainer']."""
    miss = []
    if not (str(ev.get("start_time", "")).strip() and str(ev.get("end_time", "")).strip()):
        miss.append("time")
    trainer = str(ev.get("trainer", "")).strip()
    if not trainer or trainer.upper() == "TBD":
        miss.append("trainer")
    region = str(ev.get("region", "")).strip()
    if not region or region.upper() == "TBD":
        miss.append("location")
    return miss


def event_view(ev):
    """Display-friendly fields for one event.

    NEVER raises on a bad date. It used to, and the caller that mattered
    (hub_modes.classes_overview) caught the exception and skipped the row — so
    one unparseable cell made a whole class disappear from every staff screen
    with nothing logged and no count to notice. A class with a broken date is
    exactly the class somebody needs to see.
    """
    from src.class_address import resolve
    d = safe_date(ev.get("event_date"))
    miss = missing_info(ev)
    trainer = str(ev.get("trainer", "")).strip()
    return {
        "found": True,
        "bad_date": d is None,
        "event_id": ev["event_id"],
        # branch location from the table, unless a floating address overrides it
        "class_address": resolve(ev["event_id"], ev.get("event_location", "")),
        "topic": ev.get("topic", ""),
        "region": ev.get("region", ""),
        "state": ev.get("state", ""),
        "branch": ev.get("branch", ""),
        "track": ev.get("track", ""),
        "trainer": "" if trainer.upper() == "TBD" else trainer,
        "timezone": ev.get("timezone", ""),
        "host_label": ev.get("host_label", ""),
        "event_location": ev.get("event_location", ""),
        "notes": ev.get("notes", ""),
        # normalised, so every downstream `date < today` compares like with like
        "event_date": d.isoformat() if d else "",
        "event_date_raw": str(ev.get("event_date", "")).strip(),
        "weekday_display": _weekday_upper(ev).title(),
        "date_display": f"{MONTHS[d.month - 1]} {d.day}, {d.year}" if d else "Date unreadable",
        "date_short": f"{MONTHS[d.month - 1][:3]} {d.day}" if d else "??",
        "time_display": time_display(ev),
        "date_info": build_date_info(ev),
        "needs_info": bool(miss),
        "missing": miss,
    }


def is_active(ev):
    return str(ev.get("active")).strip().lower() not in ("false", "0", "no", "")


def event_cache_row(ev):
    """Fields persisted to the DB events cache (referential integrity / reporting)."""
    return {
        "event_id": ev["event_id"],
        "topic": ev.get("topic", ""),
        "region": ev.get("region", ""),
        "branch": ev.get("branch", ""),
        "event_date": str(ev.get("event_date", "")),
        "start_time": str(ev.get("start_time", "")),
        "end_time": str(ev.get("end_time", "")),
        "event_location": ev.get("event_location", ""),
        "host_label": ev.get("host_label", ""),
        "capacity": str(ev.get("capacity", "")),
    }


def list_events(include_past=False):
    """Active classes as display views, newest-first by id.

    Dealers only ever see today's and future classes. The team needs the
    finished ones too — /admin and /manage are gone, so the main list is the
    only place an FSR can reach a past class to grade it. That is what
    include_past opens up, and it is never set from the public feed.

    Cancelled classes (active=false) stay hidden either way.
    """
    events, _, _ = load_catalog()
    today = str(datetime.now().date())
    out = []
    for ev in events.values():
        if not is_active(ev):
            continue
        d = safe_date(ev.get("event_date"))
        if d is None:
            continue          # a class we cannot date is never offered to dealers
        if not include_past and d.isoformat() < today:
            continue
        try:
            v = event_view(ev)
        except (ValueError, TypeError, KeyError):
            continue
        out.append(v)
    out.sort(key=lambda v: v["event_id"])
    return out


def public_events():
    """The dealer-facing feed: today and later, never the finished classes."""
    return list_events(include_past=False)
