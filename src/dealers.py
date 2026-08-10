"""Program-dealer directory read live from data/dealers.xlsx.

Sheet layout (John's export, sheet "Results"): two banner rows, then headers on
row 3 — Customer ID | Customer Name | Sales Rep Name (Cust) — data from row 4.
Like the events catalog, the file is re-read when its mtime changes, so John can
drop in a fresh export without restarting the server.

Sales rep arrives as "Last, First" and is normalized to "First Last" so it can
land directly in the Territory Manager column of the export.
"""
from openpyxl import load_workbook

from config import DEALERS_XLSX

_cache = {"mtime": None, "by_id": {}, "by_name": {}, "directory": []}

# The dealer sheet uses nicknames; the employee list (and the TM access keys)
# use legal names. Normalize at the source so every stored Territory Manager
# matches one canonical spelling everywhere.
REP_ALIASES = {
    "Drew Johnson": "Andrew Johnson",
    "Matt Barber": "Matthew Barber",
    "Matt Griffin": "Matthew Griffin",
    "Matt Smith": "Matthew Smith",
    "Mike Bacon": "Michael Bacon",
}


def _flip_name(rep):
    rep = str(rep or "").strip()
    if "," in rep:
        last, _, first = rep.partition(",")
        rep = f"{first.strip()} {last.strip()}".strip()
    return REP_ALIASES.get(rep, rep)


def _norm(name):
    return " ".join(str(name or "").split()).casefold()


def _load():
    mtime = DEALERS_XLSX.stat().st_mtime if DEALERS_XLSX.exists() else None
    if mtime == _cache["mtime"]:
        return
    by_id, by_name, directory = {}, {}, []
    if mtime is not None:
        wb = load_workbook(DEALERS_XLSX, read_only=True, data_only=True)
        ws = wb.active
        header_seen = False
        for row in ws.iter_rows(values_only=True):
            cid, name, rep = (list(row) + [None, None, None])[:3]
            if not header_seen:
                header_seen = str(cid or "").strip() == "Customer ID"
                continue
            name = str(name or "").strip()
            if not name or cid is None:
                continue
            dealer = {
                "customer_id": str(cid).strip(),
                "company_name": name,
                "territory_manager": _flip_name(rep),
            }
            by_id[dealer["customer_id"]] = dealer
            by_name[_norm(name)] = dealer
            directory.append({"customer_id": dealer["customer_id"], "company_name": name})
        wb.close()
        directory.sort(key=lambda d: d["company_name"].casefold())
    _cache.update(mtime=mtime, by_id=by_id, by_name=by_name, directory=directory)


def dealer_directory():
    """[{customer_id, company_name}, ...] for the form typeahead (no TM exposed)."""
    _load()
    return _cache["directory"]


def find_dealer(customer_id="", company_name=""):
    """Resolve a dealer by id first, then exact (case/space-insensitive) name."""
    _load()
    dealer = _cache["by_id"].get(str(customer_id or "").strip())
    if not dealer and company_name:
        dealer = _cache["by_name"].get(_norm(company_name))
    return dealer
