"""Edits & Updates hub — the write side of the class catalog.

catalog.py stays read-only (it derives display fields). This module is the one
place that WRITES back to data/events.xlsx: editing a class, adding/removing
seats (capacity), cancelling / reinstating, and creating a new class. John can
still open the xlsx by hand — this just gives him a friendly screen to do it.

Every write reloads the workbook, edits the matching row in the 'events' sheet,
and saves. Seats-taken comes from the live registrations DB (passed in as repo)
so capacity is never set below what's already booked.
"""
import os
import re
import shutil
import tempfile
import threading
from datetime import datetime

from openpyxl import load_workbook

from config import EVENTS_XLSX, MONTHS
from src.catalog import load_catalog, missing_info, time_display
from src.class_schedule import (parse_event_date, refuse_bad_times, refuse_past_date,
                               to_hhmm)

# Several people can edit at once (two admins on two class pages, or an admin
# while a tool runs). openpyxl rewrites the WHOLE workbook on save, so two
# overlapping saves interleave and leave a corrupt zip — the catalog is then
# unreadable and every page 500s. This lock serialises catalog writes inside the
# process; _atomic_save makes each write all-or-nothing on disk.
_CATALOG_LOCK = threading.RLock()


def _atomic_save(wb, path=EVENTS_XLSX):
    """Write to a temp file in the same directory, then rename over the target.

    rename() is atomic on POSIX, so a reader either sees the whole old file or
    the whole new one — never a half-written zip. A crash mid-write leaves the
    original untouched.
    """
    path = str(path)
    fd, tmp = tempfile.mkstemp(suffix=".xlsx", dir=os.path.dirname(path) or ".")
    os.close(fd)
    try:
        wb.save(tmp)
        shutil.copystat(path, tmp) if os.path.exists(path) else None
        os.replace(tmp, path)
    finally:
        if os.path.exists(tmp):
            os.unlink(tmp)

# Columns the hub is allowed to edit (header name in the 'events' sheet).
EDITABLE = ["region", "branch", "weekday", "event_date", "start_time", "end_time",
            "topic", "trainer", "event_location", "host_label", "capacity", "notes", "active"]


def _headers(ws):
    return [c.value for c in ws[1]]


def _find_row(ws, headers, event_id):
    col = headers.index("event_id") + 1
    for r in range(2, ws.max_row + 1):
        if str(ws.cell(row=r, column=col).value or "").strip() == event_id:
            return r
    return None


def _seats_taken(repo, event_id):
    try:
        return int(repo.seats_taken(event_id))
    except Exception:  # noqa: BLE001
        return 0


def list_classes(repo):
    """All classes (active + cancelled, past + future) with edit fields + seat counts."""
    from src.class_address import all_addresses
    events, branches, _ = load_catalog()
    addresses = all_addresses()          # one read, not one per class
    out = []
    today = str(datetime.now().date())
    for ev in events.values():
        eid = ev["event_id"]
        taken = _seats_taken(repo, eid)
        try:
            cap = int(float(ev.get("capacity") or 0))
        except (ValueError, TypeError):
            cap = 0
        date = str(ev.get("event_date", ""))[:10]
        active = str(ev.get("active")).strip().lower() not in ("false", "0", "no", "")
        miss = missing_info(ev)
        out.append({
            "event_id": eid,
            "topic": ev.get("topic", ""),
            "region": ev.get("region", ""),
            "branch": ev.get("branch", ""),
            "trainer": str(ev.get("trainer", "")).strip(),
            "weekday": ev.get("weekday", ""),
            "event_date": date,
            "start_time": str(ev.get("start_time", "")),
            "end_time": str(ev.get("end_time", "")),
            "time_display": time_display(ev),   # one formatter, zone included
            "event_location": ev.get("event_location", ""),
            # floating override; blank means "use the branch location above"
            "class_address": str(addresses.get(eid, "")).strip(),
            "address_display": str(addresses.get(eid, "")).strip() or ev.get("event_location", ""),
            "host_label": ev.get("host_label", ""),
            "notes": ev.get("notes", ""),
            "capacity": cap,
            "seats_taken": taken,
            "seats_left": max(0, cap - taken) if cap else None,
            "active": active,
            "is_past": bool(date) and date < today,
            "missing": miss,
            "needs_info": bool(miss),
        })
    # Control-panel order: upcoming/live first (soonest first), finished last.
    out.sort(key=lambda v: (v["is_past"], v["event_date"], v["event_id"]))
    return {"classes": out, "branches": branches}


def _derive_weekday(iso):
    try:
        return datetime.strptime(iso, "%Y-%m-%d").strftime("%A").upper()
    except (ValueError, TypeError):
        return ""


def has_grade_history(repo, event_id):
    """True when anyone on this class has ever been graded.

    This is what makes a past date legitimate: a class with grades on it
    demonstrably happened, so restoring the day it happened on is a correction,
    not a fabrication. A clean upcoming class has no such evidence.
    """
    try:
        return any(r.get("attended") is not None for r in repo.class_grades(event_id))
    except Exception:  # noqa: BLE001
        return False


def update_class(repo, event_id, fields, allow_past_restore=False):
    """Patch any subset of EDITABLE fields on one class. Returns {ok, message, class}.

    allow_past_restore is the escape hatch for the one-way door: moving a
    finished class forward is allowed (it fixes a typo), which used to mean the
    move could never be undone, because putting the real date back is by
    definition writing a past date. It is granted ONLY by an explicit admin
    confirmation, or automatically when the class already carries grades.

    Serialised: concurrent admins editing the same (or any) class cannot
    interleave workbook writes.
    """
    with _CATALOG_LOCK:
        return _update_class(repo, event_id, fields, allow_past_restore)


def _update_class(repo, event_id, fields, allow_past_restore=False):
    event_id = str(event_id).strip()
    if not event_id:
        return {"ok": False, "error": "Missing event_id."}

    fields = dict(fields or {})
    # class_address is a floating overlay, NOT a table column — it is written to
    # data/class_address.json and deliberately never reaches the workbook.
    changed_address = []
    if "class_address" in fields:
        from src.class_address import set_address
        res = set_address(event_id, fields.pop("class_address"))
        if not res.get("ok"):
            return res
        changed_address.append("class_address")

    if not fields:                       # address-only edit: never open the workbook
        if not changed_address:          # ...and nothing at all was sent
            return {"ok": False, "error": "No changes to save."}
        fresh = next((c for c in list_classes(repo)["classes"] if c["event_id"] == event_id), None)
        return {"ok": True, "message": f"Saved {len(changed_address)} change(s).",
                "changed": changed_address, "class": fresh}

    wb = load_workbook(EVENTS_XLSX)
    ws = wb["events"]
    headers = _headers(ws)
    row = _find_row(ws, headers, event_id)
    if not row:
        wb.close()
        return {"ok": False, "error": f"Class '{event_id}' not found."}

    # ---- schedule guards -------------------------------------------------
    # Run BEFORE a single cell is touched, so a refusal leaves the workbook and
    # the file on disk exactly as they were. src.class_schedule owns the rules
    # (and the same clock the past/upcoming split uses everywhere else).
    if "event_date" in fields:
        new_date = parse_event_date(fields["event_date"])
        if not new_date:
            wb.close()
            return {"ok": False, "error": "Date must be YYYY-MM-DD."}
        err = refuse_past_date(new_date)
        if err:
            # The two ways a past date is legitimate. Both mean "this class
            # really happened then" rather than "I am backdating a live class".
            restore = bool(allow_past_restore) or has_grade_history(repo, event_id)
            if not restore:
                wb.close()
                return {"ok": False, "error": err,
                        # tells the browser an escape hatch exists for this class
                        "can_restore": False, "past_date": True}
            from src.audit import log
            log("admin", "class.restore_past_date", event_id,
                {"to": new_date.isoformat(),
                 "reason": "explicit" if allow_past_restore else "has grade history"})
        fields["event_date"] = new_date.isoformat()      # normalised on the way in

    if "start_time" in fields or "end_time" in fields:
        # merge the patch over the row's CURRENT times: changing only the end
        # time still has to be checked against the start time already on file.
        def _cell(name):
            return ws.cell(row=row, column=headers.index(name) + 1).value \
                if name in headers else ""
        start = fields.get("start_time", _cell("start_time"))
        end = fields.get("end_time", _cell("end_time"))
        err = refuse_bad_times(start, end)
        if err:
            wb.close()
            return {"ok": False, "error": err}

    changed = []
    for key, raw in fields.items():
        if key not in EDITABLE or key not in headers:
            continue
        col = headers.index(key) + 1

        if key == "capacity":
            try:
                new_cap = int(float(raw))
            except (ValueError, TypeError):
                wb.close()
                return {"ok": False, "error": "Capacity must be a whole number."}
            taken = _seats_taken(repo, event_id)
            if new_cap < taken:
                wb.close()
                return {"ok": False, "error": (
                    f"Can't set capacity to {new_cap} — {taken} seat(s) are already "
                    f"booked. Set it to {taken} or higher.")}
            value = new_cap
        elif key == "active":
            value = bool(raw) if isinstance(raw, bool) else \
                str(raw).strip().lower() not in ("false", "0", "no", "")
        elif key in ("start_time", "end_time"):
            # validated above; store canonical HH:MM so time_display can read it
            value = "" if str(raw or "").strip() == "" else to_hhmm(raw)
        elif key == "event_date":
            value = str(raw).strip()
            # keep weekday in sync with a new date, unless caller set it explicitly
            if "weekday" not in fields and "weekday" in headers:
                wd = _derive_weekday(value)
                if wd:
                    ws.cell(row=row, column=headers.index("weekday") + 1).value = wd
                    changed.append("weekday")
        else:
            value = str(raw).strip()

        ws.cell(row=row, column=col).value = value
        changed.append(key)

    _atomic_save(wb)
    wb.close()
    changed += changed_address
    fresh = next((c for c in list_classes(repo)["classes"] if c["event_id"] == event_id), None)
    return {"ok": True, "message": f"Saved {len(set(changed))} change(s).",
            "changed": sorted(set(changed)), "class": fresh}


def adjust_seats(repo, event_id, delta):
    """Add (or remove) seats relative to current capacity."""
    try:
        delta = int(delta)
    except (ValueError, TypeError):
        return {"ok": False, "error": "Seat change must be a whole number."}
    current = next((c for c in list_classes(repo)["classes"] if c["event_id"] == str(event_id).strip()), None)
    if not current:
        return {"ok": False, "error": f"Class '{event_id}' not found."}
    return update_class(repo, event_id, {"capacity": max(0, current["capacity"] + delta)})


def set_active(repo, event_id, active):
    """Cancel (active=False) or reinstate (active=True) a class."""
    result = update_class(repo, event_id, {"active": active})
    if result.get("ok"):
        result["message"] = "Class reinstated — dealers can register again." if active \
            else "Class cancelled — it's now hidden from dealers."
    return result


def _slugify(text):
    return re.sub(r"[^a-z0-9]+", "-", str(text).lower()).strip("-")


def create_class(repo, fields):
    """Append a brand-new class. region + topic + event_date are required."""
    with _CATALOG_LOCK:
        return _create_class(repo, fields)


def _create_class(repo, fields):
    region = str(fields.get("region", "")).strip()
    topic = str(fields.get("topic", "")).strip()
    date = str(fields.get("event_date", "")).strip()
    if not (region and topic and date):
        return {"ok": False, "error": "Region, topic and date are required to add a class."}

    # Same guard as the edit path, before the id is even minted: a class that
    # was born in the past is invisible to dealers and lands straight in the
    # FSR grading queue, which is never what "add a class" meant.
    parsed = parse_event_date(date)
    if not parsed:
        return {"ok": False, "error": "Date must be YYYY-MM-DD."}
    err = refuse_past_date(parsed)
    if err:
        return {"ok": False, "error": err}
    date = parsed.isoformat()
    fields = {**fields, "event_date": date}

    # build a readable, unique event_id: region-firstword-YYYY-MM-DD
    first = _slugify(topic.split()[0] if topic.split() else "class")
    base = f"{_slugify(region)}-{first}-{date}"
    events, _, _ = load_catalog()
    eid, n = base, 2
    while eid in events:
        eid, n = f"{base}-{n}", n + 1

    # TEMPLATE: a new class inherits the region's settled facts (state, timezone,
    # branch, venue, host line) from the classes already running there. Without
    # this a new class lands half-blank — no state on the location line, no
    # timezone, no branch — and the class page can't show what it's capable of.
    sibling = _region_template(events, region, str(fields.get("branch", "")).strip())
    new_branch = False

    def pick(key, fallback=""):
        """caller's value > the region's existing pattern > fallback"""
        given = str(fields.get(key, "") or "").strip()
        return given or str(sibling.get(key, "") or "").strip() or fallback

    wb = load_workbook(EVENTS_XLSX)
    ws = wb["events"]
    headers = _headers(ws)
    defaults = {
        "active": True,
        "event_id": eid,
        "region": region,
        "branch": pick("branch"),
        "state": pick("state"),
        "weekday": _derive_weekday(date),
        "event_date": date,
        "start_time": pick("start_time", "09:00"),
        "end_time": pick("end_time", "14:00"),
        "timezone": pick("timezone"),
        "topic": topic,
        "track": pick("track"),
        "trainer": pick("trainer", "TBD"),
        "alt_trainer": str(fields.get("alt_trainer", "") or "").strip(),
        "event_location": pick("event_location", f"{region.upper()} BRANCH"),
        "host_label": pick("host_label", f"Hosted by M&A Supply — {region}"),
        "capacity": int(float(fields.get("capacity") or sibling.get("capacity") or 18)),
        "notes": str(fields.get("notes", "") or "").strip(),
        "status": "",
    }

    # Times can arrive from three places (caller > sibling class > 09:00/14:00),
    # so the pair is checked AFTER it's resolved — an inherited or defaulted
    # pair has to be as legal as a typed one.
    err = refuse_bad_times(defaults["start_time"], defaults["end_time"])
    if err:
        wb.close()
        return {"ok": False, "error": err}
    defaults["start_time"] = to_hhmm(defaults["start_time"])
    defaults["end_time"] = to_hhmm(defaults["end_time"])

    ws.append([defaults.get(h, "") for h in headers])

    # A brand-new branch must also land in the `branches` sheet. That sheet is
    # what fills the dealer's "Your M&A Branch" dropdown AND what
    # registrations.py validates against — without this row the class exists
    # but nobody can register for it.
    branch_name = defaults.get("branch", "").strip()
    if branch_name and "branches" in wb.sheetnames:
        bs = wb["branches"]
        known = {str(bs.cell(row=r, column=1).value or "").strip()
                 for r in range(2, bs.max_row + 1)}
        if branch_name not in known:
            bs.append([branch_name, ""])          # TM column filled in later
            new_branch = True

    _atomic_save(wb)
    wb.close()

    fresh = next((c for c in list_classes(repo)["classes"] if c["event_id"] == eid), None)
    missing = (fresh or {}).get("missing") or []
    msg = f"Added “{topic}” in {region}."
    if new_branch:
        msg += f" New branch “{branch_name}” added — dealers can now pick it."
    if sibling:
        msg += f" Branch, state, timezone and venue copied from {region}'s other classes."
    if missing:
        msg += f" Still needs: {', '.join(missing)}."
    return {"ok": True, "message": msg, "class": fresh, "inherited": bool(sibling),
            "missing": missing}


def _region_template(events, region, branch=""):
    """The most recent active class in the same branch (or failing that, the same
    region) — used as the shape a new class is cast from. Returns {} when this is
    the first class somewhere new, and the plain defaults apply."""
    reg = region.strip().lower()
    br = branch.strip().lower()
    pool = [e for e in events.values()
            if str(e.get("active")).strip().lower() not in ("false", "0", "no")]
    same_branch = [e for e in pool if br and str(e.get("branch", "")).strip().lower() == br]
    same_region = [e for e in pool if str(e.get("region", "")).strip().lower() == reg]
    best = same_branch or same_region
    if not best:
        return {}
    # newest first — the latest class reflects how the region is running today
    return sorted(best, key=lambda e: str(e.get("event_date", "")), reverse=True)[0]
