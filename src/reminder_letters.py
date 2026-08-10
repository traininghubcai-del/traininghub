"""Per-student reminder letters, rendered to PDF in memory.

Today there is no mail system. The workflow is: staff opens one letter per
student, prints or saves it, and hands it over — so *generating* a letter is the
send. That is why generate() writes to the SAME outbox ledger that
src/email_campaign.py appends to (see src/reminders.py): the "total reminders
sent" and "last reminder sent on" columns in ADMIN-VIEW are driven by that file,
so a print run shows up there immediately and correctly.

When a provider is wired in later, the context built here is what the email
template should render too — one letter definition, two delivery paths.

Nothing is written to disk except the ledger row. PDFs are built per request and
streamed; no file ever lands under data/.
"""
import io
from datetime import datetime

from config import CAMPAIGN_OUTBOX_XLSX, EMAIL_CAMPAIGN_DIR

# ledger column order, mirrored from email_campaign.OUTBOX_COLUMNS
_OUTBOX_COLS = ["sent_at", "simulated_today", "stage_days", "send_on", "reg_id",
                "contact_email", "company_name", "event_id", "topic",
                "class_date", "subject", "html_file", "status"]

NAVY = (10 / 255, 37 / 255, 64 / 255)
BLUE = (29 / 255, 111 / 255, 194 / 255)
MUTED = (86 / 255, 105 / 255, 122 / 255)
LINE = (217 / 255, 226 / 255, 236 / 255)
GOLD = (232 / 255, 179 / 255, 44 / 255)
OK = (31 / 255, 138 / 255, 76 / 255)


def days_until(event_date):
    """Whole days from today to the class. Negative once it's past."""
    try:
        d = datetime.strptime(str(event_date)[:10], "%Y-%m-%d").date()
    except (ValueError, TypeError):
        return None
    return (d - datetime.now().date()).days


def countdown_line(days):
    if days is None:
        return "Your upcoming M&A Supply training"
    if days > 1:
        return f"You are {days} days away from your class"
    if days == 1:
        return "Your class is tomorrow"
    if days == 0:
        return "Your class is today"
    return "Class summary"


def build_context(student, event, extra=None):
    """Everything one letter needs, from data we already hold.

    student — a row from repo.class_grades()
    event   — catalog.event_view() output
    """
    days = days_until(event.get("event_date"))
    ctx = {
        "student_name": (student.get("name") or "").strip() or "Attendee",
        "role": (student.get("role") or "").strip(),
        "company_name": (student.get("company_name") or "").strip(),
        "contact_email": (student.get("contact_email") or "").strip(),
        "branch": (student.get("branch") or "").strip(),
        "topic": event.get("topic", ""),
        "track": event.get("track", ""),
        "weekday_display": event.get("weekday_display", ""),
        "date_display": event.get("date_display", ""),
        "event_date": str(event.get("event_date", ""))[:10],
        "time_display": event.get("time_display", ""),
        "timezone": event.get("timezone", ""),
        "location": event.get("class_address") or event.get("event_location", ""),
        "region": event.get("region", ""),
        "state": event.get("state", ""),
        "host_label": event.get("host_label", ""),
        "trainer": event.get("trainer", ""),
        "event_id": event.get("event_id", ""),
        "days_until": days,
        "countdown": countdown_line(days),
    }
    # real contacts for this class's branch — never an invented number
    try:
        from src.contacts import support_for
        sup = support_for(event)
        ctx.update({"branch_label": sup["branch_label"], "branch_phone": sup["phone"],
                    "regional_manager": sup["regional_manager"], "territory": sup["territory"]})
    except Exception:  # noqa: BLE001
        pass
    ctx.update(extra or {})
    return ctx


# ------------------------------------------------------------------ render --

def render_pdf(ctx):
    """One letter, one page, returned as PDF bytes."""
    return render_many([ctx], title=f"Reminder — {ctx['student_name']} — {ctx['topic']}")


def render_many(contexts, title="Reminder letters"):
    """Every letter in ONE document, one page each. This is what "open all"
    uses: a single tab the browser can never block, and one print job for the
    whole class instead of N."""
    from reportlab.lib.pagesizes import LETTER
    from reportlab.pdfgen import canvas

    buf = io.BytesIO()
    c = canvas.Canvas(buf, pagesize=LETTER)
    c.setTitle(title)
    for ctx in contexts:
        _draw_letter(c, ctx)
        c.showPage()
    c.save()
    return buf.getvalue()


def _draw_letter(c, ctx):
    """One complete, upbeat letter on one page: what's confirmed, when and where,
    what they get, what to bring, and exactly who to call."""
    from reportlab.lib.pagesizes import LETTER
    from reportlab.lib.units import inch

    W, H = LETTER
    m = 0.72 * inch
    inner = W - 2 * m

    # ---- masthead -------------------------------------------------------
    c.setFillColorRGB(*NAVY)
    c.rect(0, H - 1.32 * inch, W, 1.32 * inch, stroke=0, fill=1)
    c.setFillColorRGB(*BLUE)
    c.rect(0, H - 1.32 * inch, W, 5, stroke=0, fill=1)
    c.setFillColorRGB(1, 1, 1)
    c.setFont("Helvetica-Bold", 19)
    c.drawString(m, H - 0.62 * inch, "M&A SUPPLY COMPANY")
    c.setFont("Helvetica", 9.5)
    c.setFillColorRGB(0.72, 0.82, 0.92)
    c.drawString(m, H - 0.86 * inch, "TRAINING HUB  ·  PROFESSIONAL HVAC TRAINING")
    # countdown chip
    chip = ctx["countdown"].upper()
    c.setFont("Helvetica-Bold", 9.5)
    cw = c.stringWidth(chip, "Helvetica-Bold", 9.5) + 18
    c.setFillColorRGB(*GOLD)
    c.rect(W - m - cw, H - 0.95 * inch, cw, 0.28 * inch, stroke=0, fill=1)
    c.setFillColorRGB(0.16, 0.11, 0.0)
    c.drawCentredString(W - m - cw / 2, H - 0.87 * inch, chip)

    y = H - 1.32 * inch - 0.44 * inch

    # ---- headline + greeting -------------------------------------------
    c.setFillColorRGB(*OK)
    c.setFont("Helvetica-Bold", 10)
    c.drawString(m, y, "YOUR SEAT IS CONFIRMED")
    y -= 0.30 * inch
    c.setFillColorRGB(*NAVY)
    c.setFont("Helvetica-Bold", 16)
    c.drawString(m, y, f"Dear {ctx['student_name']},")
    y -= 0.26 * inch
    y = _paragraph(c, _intro_text(ctx), m, y, inner, size=10.3, leading=14.2,
                   color=(0.09, 0.13, 0.18))
    y -= 0.12 * inch

    # ---- class detail box ----------------------------------------------
    rows = [r for r in (
        ("Class", ctx["topic"]),
        ("Level", ctx["track"]),
        ("Date", f"{ctx['weekday_display']}, {ctx['date_display']}".strip(", ")),
        ("Time", f"{ctx['time_display']} {ctx['timezone']}".strip()),
        ("Location", ctx["location"]),
        ("Branch", ctx["branch"] or ctx["region"]),
        ("Instructor", ctx["trainer"]),
    ) if str(r[1]).strip()]

    row_h = 0.235 * inch
    box_h = row_h * len(rows) + 0.22 * inch
    c.setFillColorRGB(0.965, 0.978, 0.992)
    c.setStrokeColorRGB(*LINE)
    c.rect(m, y - box_h, inner, box_h, stroke=1, fill=1)
    c.setFillColorRGB(*BLUE)
    c.rect(m, y - box_h, 4, box_h, stroke=0, fill=1)
    ty = y - 0.24 * inch
    for label, value in rows:
        c.setFillColorRGB(*MUTED)
        c.setFont("Helvetica-Bold", 7.4)
        c.drawString(m + 0.22 * inch, ty, label.upper())
        c.setFillColorRGB(*NAVY)
        c.setFont("Helvetica-Bold", 10.2)
        c.drawString(m + 1.18 * inch, ty, str(value)[:64])
        ty -= row_h
    y -= box_h + 0.26 * inch

    # ---- what you get ---------------------------------------------------
    y = _heading(c, "WHAT'S INCLUDED", m, y)
    perks = ["No cost to attend", "Lunch provided", "Hands-on with real equipment",
             "Free parking at the branch", "Counts toward your training record"]
    y = _bullets(c, perks, m, y, inner, cols=2)
    y -= 0.14 * inch

    # ---- what to bring --------------------------------------------------
    y = _heading(c, "WHAT TO BRING", m, y)
    y = _bullets(c, ["Your own hand tools", "PPE (safety glasses, gloves, boots)",
                     "Arrive 15 minutes early to sign in"], m, y, inner, cols=2)
    y -= 0.16 * inch

    # ---- questions / contact -------------------------------------------
    qh = 0.86 * inch
    c.setFillColorRGB(0.996, 0.972, 0.937)
    c.setStrokeColorRGB(*GOLD)
    c.rect(m, y - qh, inner, qh, stroke=1, fill=1)
    c.setFillColorRGB(0.55, 0.36, 0.03)
    c.setFont("Helvetica-Bold", 9.6)
    c.drawString(m + 0.22 * inch, y - 0.26 * inch, "QUESTIONS ABOUT THIS CLASS?")
    c.setFillColorRGB(0.09, 0.13, 0.18)
    c.setFont("Helvetica", 9.4)
    bits = []
    if ctx.get("branch_phone"):
        bits.append(f"Call the {ctx.get('branch_label') or 'M&A'} branch at {ctx['branch_phone']}")
    elif ctx.get("branch_label"):
        bits.append(f"Call your {ctx['branch_label']} branch")
    if ctx["trainer"]:
        bits.append(f"or ask your instructor {ctx['trainer']}")
    c.drawString(m + 0.22 * inch, y - 0.48 * inch, (", ".join(bits) or "Contact your M&A branch") + ".")
    second = "Your M&A Territory Manager can answer anything about your account or seats."
    if ctx.get("regional_manager"):
        second = (f"Regional Manager {ctx['regional_manager']} covers "
                  f"{ctx.get('territory','your area')} — your TM can also help with seats.")
    c.drawString(m + 0.22 * inch, y - 0.67 * inch, second)
    y -= qh + 0.22 * inch

    # ---- cancellation ---------------------------------------------------
    c.setFillColorRGB(*MUTED)
    c.setFont("Helvetica-Oblique", 8.8)
    y = _paragraph(c, ("Can't make it? Let us know at least 24 hours ahead so the seat can go to "
                       "another dealer. Seats are limited and these classes fill up."),
                   m, y, inner, size=8.8, leading=12, color=MUTED)

    # ---- footer ---------------------------------------------------------
    c.setStrokeColorRGB(*LINE)
    c.line(m, m + 0.40 * inch, W - m, m + 0.40 * inch)
    c.setFillColorRGB(*MUTED)
    c.setFont("Helvetica", 8.2)
    who = ", ".join(x for x in (ctx["company_name"], ctx["role"]) if x)
    c.drawString(m, m + 0.22 * inch,
                 f"Registered through {who}." if who else (ctx["host_label"] or "M&A Supply Company Inc."))
    c.drawRightString(W - m, m + 0.22 * inch, "masupplycompany.com")


def _intro_text(ctx):
    """The opener. A class is worth showing up for — this should read like it."""
    d = ctx["days_until"]
    when = f"{ctx['weekday_display']}, {ctx['date_display']}".strip(", ")
    topic, where = ctx["topic"], ctx["location"]
    if d is not None and d > 1:
        return (f"You're {d} days out from {topic} — and your seat is locked in. "
                f"We'll see you {when} at {where}. This one is hands-on: real equipment, "
                f"real troubleshooting, and an instructor who does this work every day. "
                f"Come ready to get your hands dirty.")
    if d == 1:
        return (f"Tomorrow is the day — {topic} at {where}. Your seat is confirmed. "
                f"Bring your tools and come ready to work through real equipment with the crew.")
    if d == 0:
        return (f"Today's the day — {topic} at {where}. Your seat is confirmed and we're "
                f"ready for you. Doors open 15 minutes before we start.")
    return f"Here are your details for {topic} on {when} at {where}."


def _heading(c, text, x, y):
    """Small blue section label. Returns the y below it."""
    from reportlab.lib.units import inch
    c.setFillColorRGB(*BLUE)
    c.setFont("Helvetica-Bold", 8.6)
    c.drawString(x, y, text)
    return y - 0.20 * inch


def _bullets(c, items, x, y, width, cols=2):
    """Compact two-column checklist."""
    from reportlab.lib.units import inch
    colw = width / cols
    rows = (len(items) + cols - 1) // cols
    for i, text in enumerate(items):
        col, row = i % cols, i // cols
        bx = x + col * colw
        by = y - row * 0.20 * inch
        c.setFillColorRGB(*OK)
        c.setFont("Helvetica-Bold", 9.5)
        c.drawString(bx, by, "\u2022")
        c.setFillColorRGB(0.09, 0.13, 0.18)
        c.setFont("Helvetica", 9.4)
        c.drawString(bx + 0.13 * inch, by, text)
    return y - rows * 0.20 * inch


def _paragraph(c, text, x, y, width, size=11, leading=15, color=(0, 0, 0)):
    """Word-wrap a paragraph. Returns the y position below it."""
    from reportlab.pdfbase.pdfmetrics import stringWidth
    c.setFillColorRGB(*color)
    c.setFont("Helvetica", size)
    words, line = str(text).split(), ""
    for w in words:
        probe = f"{line} {w}".strip()
        if stringWidth(probe, "Helvetica", size) <= width:
            line = probe
        else:
            c.drawString(x, y, line)
            y -= leading
            line = w
    if line:
        c.drawString(x, y, line)
        y -= leading
    return y


# ------------------------------------------------------------------ ledger --

def log_generated(students, event, stage_days=None):
    """Record one ledger row per letter, so ADMIN-VIEW's reminder columns show
    this print run. Same file + column order the email sender uses, so the two
    can never disagree about what went out."""
    from openpyxl import Workbook, load_workbook

    EMAIL_CAMPAIGN_DIR.mkdir(parents=True, exist_ok=True)
    if CAMPAIGN_OUTBOX_XLSX.exists():
        wb = load_workbook(CAMPAIGN_OUTBOX_XLSX)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "outbox"
        ws.append(_OUTBOX_COLS)

    now = datetime.now()
    sent_at = now.isoformat(timespec="seconds")
    today = str(now.date())
    days = days_until(event.get("event_date")) if stage_days is None else stage_days
    eid = event.get("event_id", "")

    for s in students:
        ws.append([
            sent_at, today, days if days is not None else "", today,
            s.get("attendee_id", ""), s.get("contact_email", ""),
            s.get("company_name", ""), eid, event.get("topic", ""),
            str(event.get("event_date", ""))[:10],
            f"Reminder letter — {event.get('topic','')}",
            "",                                   # no html file: this is a PDF letter
            "PRINTED — reminder letter generated as PDF",
        ])
    wb.save(CAMPAIGN_OUTBOX_XLSX)
    wb.close()
    return {"logged": len(students), "sent_at": sent_at, "days_until": days}
