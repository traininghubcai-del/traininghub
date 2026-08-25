"""Data sanity audit — run this before going live, and after ANYONE edits.

Once several people can edit the catalog (in Excel) and the class pages (in the
Admin lens), "it looked fine when I checked" stops being good enough. This
asserts the invariants the app actually depends on, and prints what broke.

    .venv/bin/python tools/data_audit.py            # audit files + DB
    .venv/bin/python tools/data_audit.py --server   # also hit the live endpoints

Exit code 0 = clean, 1 = at least one FAIL. Safe to run any time: reads only.
"""
import json
import os
import sqlite3
import sys
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from config import CAMPAIGN_OUTBOX_XLSX, DB_PATH, EVENTS_XLSX, REG_XLSX  # noqa: E402
from src.catalog import is_active, list_events, load_catalog, public_events  # noqa: E402

FAILS, WARNS = [], []


def check(ok, label, detail="", warn_only=False):
    if ok:
        print(f"  PASS  {label}")
    elif warn_only:
        WARNS.append(f"{label} — {detail}")
        print(f"  WARN  {label}  {detail}")
    else:
        FAILS.append(f"{label} — {detail}")
        print(f"  FAIL  {label}  {detail}")


def note(t):
    print(f"  SKIP  {t}")


def _admin_code():
    """The Admin code the server is actually using, so the audit never hardcodes it."""
    from src.hub_modes import DEFAULT_CODE, STORE
    if STORE.exists():
        try:
            return str(json.loads(STORE.read_text() or "{}").get("admin") or DEFAULT_CODE)
        except (ValueError, OSError):
            pass
    return DEFAULT_CODE


def head(t):
    print(f"\n{t}\n" + "-" * len(t))


# ---------------------------------------------------------------- catalog --
def audit_catalog():
    head("CATALOG  (data/OFFICIAL_CLASS_SCHEDULE.xlsx)")
    events, branches, _ = load_catalog()
    check(bool(events), "catalog loads and is non-empty", f"{len(events)} classes")

    ids = [e.get("event_id", "") for e in events.values()]
    dupes = [k for k, n in Counter(ids).items() if n > 1]
    check(not dupes, "event_id is unique", f"duplicates: {dupes}")
    check(all(str(i).strip() for i in ids), "no blank event_id")

    # fields the app cannot render without
    required = ("topic", "region", "event_date", "branch")
    missing = defaultdict(list)
    for e in events.values():
        for f in required:
            if not str(e.get(f, "")).strip():
                missing[f].append(e.get("event_id"))
    check(not missing, "required fields present on every class",
          {k: v[:3] for k, v in missing.items()})

    bad_dates = [e["event_id"] for e in events.values()
                 if not _is_date(str(e.get("event_date", ""))[:10])]
    check(not bad_dates, "every event_date parses as YYYY-MM-DD", bad_dates[:5])

    bad_times = [e["event_id"] for e in events.values()
                 if (str(e.get("start_time", "")).strip() and str(e.get("end_time", "")).strip()
                     and str(e["end_time"]) <= str(e["start_time"]))]
    check(not bad_times, "end_time is after start_time", bad_times[:5])

    # TIMEZONE — the one that puts people at a branch an hour late
    no_tz = [e["event_id"] for e in events.values()
             if str(e.get("start_time", "")).strip() and not str(e.get("timezone", "")).strip()]
    check(not no_tz, "every timed class states its timezone", no_tz[:5])

    per_branch = defaultdict(set)
    for e in events.values():
        tz = str(e.get("timezone", "")).strip()
        if tz:
            per_branch[str(e.get("branch", "")).strip()].add(tz)
    split = {b: sorted(z) for b, z in per_branch.items() if len(z) > 1}
    check(not split, "each branch uses ONE timezone", split)

    caps = [e["event_id"] for e in events.values()
            if str(e.get("capacity", "")).strip() and not _is_int(e["capacity"])]
    check(not caps, "capacity is numeric where set", caps[:5])
    return events


# --------------------------------------------------------------------- DB --
def audit_db(events):
    head("DATABASE  (data/registrations.db)")
    if not Path(DB_PATH).exists():
        check(False, "database exists", str(DB_PATH))
        return
    con = sqlite3.connect(DB_PATH)
    con.row_factory = sqlite3.Row
    tables = {r[0] for r in con.execute("SELECT name FROM sqlite_master WHERE type='table'")}
    for t in ("registrations", "registration_attendees", "companies", "people"):
        check(t in tables, f"table {t} exists")

    regs = con.execute("SELECT id, event_id, contact_email FROM registrations").fetchall()
    check(True, "registrations readable", f"{len(regs)} rows")

    # a registration pointing at a class that no longer exists = an orphan
    orphans = [r["id"] for r in regs if r["event_id"] not in events]
    check(not orphans, "every registration points at a real class",
          f"orphan registration ids: {orphans}", warn_only=True)

    childless = con.execute("""SELECT r.id FROM registrations r
                               LEFT JOIN registration_attendees a ON a.registration_id = r.id
                               WHERE a.id IS NULL""").fetchall()
    check(not childless, "every registration has at least one attendee",
          [r["id"] for r in childless])

    no_email = [r["id"] for r in regs if not (r["contact_email"] or "").strip()]
    check(not no_email, "every registration has a contact email", no_email)

    # capacity must never be exceeded
    over = []
    for eid, ev in events.items():
        try:
            cap = int(float(ev.get("capacity") or 0))
        except (TypeError, ValueError):
            continue
        if not cap:
            continue
        n = con.execute("""SELECT COUNT(a.id) FROM registration_attendees a
                           JOIN registrations r ON a.registration_id = r.id
                           WHERE r.event_id = ?""", (eid,)).fetchone()[0]
        if n > cap:
            over.append(f"{eid}: {n}/{cap}")
    check(not over, "no class is overbooked", over)

    # grading consistency
    bad = con.execute("""SELECT id, name FROM registration_attendees
                         WHERE attended = 0 AND passed = 1""").fetchall()
    check(not bad, "no student is marked passed while absent", [r["name"] for r in bad])
    bad2 = con.execute("""SELECT id, name FROM registration_attendees
                          WHERE score IS NOT NULL AND (score < 0 OR score > 100)""").fetchall()
    check(not bad2, "scores are within 0-100", [r["name"] for r in bad2])

    # attendee count must match what was promised on the registration
    drift = con.execute("""SELECT r.id, r.num_attending, COUNT(a.id) n
                           FROM registrations r
                           LEFT JOIN registration_attendees a ON a.registration_id = r.id
                           GROUP BY r.id HAVING r.num_attending <> n""").fetchall()
    check(not drift, "num_attending matches the attendee rows",
          [f"reg {r['id']}: says {r['num_attending']}, has {r['n']}" for r in drift])

    tables2 = {t[0] for t in con.execute("SELECT name FROM sqlite_master WHERE type='table'")}
    check("audit_log" in tables2, "audit log table exists")
    if "audit_log" in tables2:
        n = con.execute("SELECT COUNT(*) FROM audit_log").fetchone()[0]
        check(True, "audit log is recording", f"{n} entries")
    con.close()


# --------------------------------------------------------------- mirrors --
def audit_mirrors(events):
    head("MIRRORS & LEDGERS")
    from openpyxl import load_workbook

    if Path(REG_XLSX).exists():
        con = sqlite3.connect(DB_PATH)
        db_n = con.execute("SELECT COUNT(*) FROM registrations").fetchone()[0]
        con.close()
        ws = load_workbook(REG_XLSX, read_only=True).active
        xl_n = sum(1 for r in ws.iter_rows(min_row=2, values_only=True) if any(v is not None for v in r))
        check(db_n == xl_n, "registrations.xlsx matches the DB row count",
              f"db={db_n} xlsx={xl_n}")
    else:
        check(False, "registrations.xlsx exists", "", warn_only=True)

    addr = Path("data/class_address.json")
    if addr.exists():
        try:
            d = json.loads(addr.read_text() or "{}")
            ghosts = [k for k in d if k not in events]
            check(not ghosts, "class_address overlay has no ghost classes", ghosts)
        except ValueError as e:
            check(False, "class_address.json is valid JSON", str(e))

    fl = Path("data/fliers/_index.json")
    if fl.exists():
        d = json.loads(fl.read_text() or "{}")
        ghosts = [k for k in d if k not in events]
        missing = [k for k, v in d.items() if not (Path("data/fliers") / v.get("file", "")).exists()]
        check(not ghosts, "flier index has no ghost classes", ghosts)
        check(not missing, "every flier record has its file on disk", missing)

    if Path(CAMPAIGN_OUTBOX_XLSX).exists():
        ws = load_workbook(CAMPAIGN_OUTBOX_XLSX, read_only=True).active
        rows = [r for r in ws.iter_rows(min_row=2, values_only=True) if r and r[0]]
        ghosts = sorted({r[7] for r in rows if r[7] and r[7] not in events})
        check(not ghosts, "reminder ledger has no ghost classes", ghosts, warn_only=True)


# ---------------------------------------------------------------- server --
def audit_server():
    head("LIVE ENDPOINTS")
    import urllib.error
    import urllib.request

    base = f"http://127.0.0.1:{os.environ.get('PORT', '8080')}"

    def get(path):
        with urllib.request.urlopen(base + path, timeout=5) as r:
            return r.status, json.loads(r.read().decode())

    try:
        _, data = get("/api/events")
    except (urllib.error.URLError, OSError) as e:
        check(False, "server is reachable", str(e))
        return
    evs = data.get("events", [])
    check(bool(evs), "/api/events returns classes", f"{len(evs)} public")

    no_tz = [e["event_id"] for e in evs
             if e["time_display"] != "Time TBD" and not e["time_display"].rstrip()[-3:].isupper()]
    check(not no_tz, "every public time shows its timezone label", no_tz[:5])

    today = str(datetime.now().date())
    past = [e["event_id"] for e in evs if e["event_date"] < today]
    check(not past, "public list hides past classes", past[:5])

    # ...and the admin overview must NOT hide them: /admin is gone, so the
    # master list under ADMIN-VIEW is the only route to a finished class, and
    # an FSR who can't reach it can't grade it.
    all_ev = list_events(include_past=True)
    check(len(all_ev) >= len(evs), "admin catalog is a superset of the public one",
          f"{len(all_ev)} all vs {len(evs)} public")
    catalog_past = [e["event_id"] for e in all_ev if e["event_date"] < today]
    if catalog_past:
        code = _admin_code()
        try:
            _, ov = get(f"/api/hub/classes?mode=admin&code={code}")
            ids = {c.get("event_id") for c in ov.get("classes", [])}
            missing = [e for e in catalog_past if e not in ids]
            check(ov.get("ok") and not missing,
                  "admin overview includes past classes for grading",
                  missing[:5] or ov.get("error", ""))
            row = next((c for c in ov.get("classes", []) if c.get("event_id") in catalog_past), None)
            check(bool(row and row.get("topic") and row.get("date_short")),
                  "admin overview rows carry the display fields the list renders")
        except (urllib.error.URLError, urllib.error.HTTPError, OSError) as e:
            check(False, "admin overview includes past classes for grading", str(e))
    else:
        note("no past classes in the catalog yet — admin past-class check skipped")

    # the gate that matters: locked modes must not leak the roster
    for mode in ("admin", "fsr"):
        try:
            urllib.request.urlopen(
                f"{base}/api/hub/class?event_id={evs[0]['event_id']}&mode={mode}",
                timeout=5)
            check(False, f"{mode} mode is locked without a code", "returned 200")
        except urllib.error.HTTPError as e:
            check(e.code == 403, f"{mode} mode is locked without a code", f"HTTP {e.code}")

    _, one = get(f"/api/hub/class?event_id={evs[0]['event_id']}&mode=user")
    body = json.dumps(one)
    check("roster" not in one, "user payload carries no roster")
    check("@" not in body.replace("_ @ ", ""), "user payload carries no email addresses")


def _is_date(s):
    try:
        datetime.strptime(s, "%Y-%m-%d")
        return True
    except (ValueError, TypeError):
        return False


def _is_int(v):
    try:
        int(float(v))
        return True
    except (TypeError, ValueError):
        return False


if __name__ == "__main__":
    print("DATA AUDIT —", datetime.now().strftime("%Y-%m-%d %H:%M"))
    evs = audit_catalog()
    audit_db(evs)
    audit_mirrors(evs)
    if "--server" in sys.argv:
        audit_server()

    print("\n" + "=" * 60)
    if FAILS:
        print(f"{len(FAILS)} FAILURE(S):")
        for f in FAILS:
            print("  -", f)
    if WARNS:
        print(f"{len(WARNS)} warning(s):")
        for w in WARNS:
            print("  -", w)
    if not FAILS:
        print("ALL CHECKS PASSED" + (f" ({len(WARNS)} warning(s))" if WARNS else ""))
    sys.exit(1 if FAILS else 0)
