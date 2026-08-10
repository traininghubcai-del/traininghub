"""Generate data/events.xlsx — the class catalog + branch list.

One Excel file, two sheets, easy for John/team to edit. One row = one class in
one location = one QR code. The app derives the combined date_info string and
all display text from these columns, so nothing has to be re-typed.

Reset to seed:  python3 make_seed_xlsx.py   (then refresh the page; no restart)

events sheet columns (left -> right):
  active, event_id, region, branch, weekday, event_date, start_time, end_time,
  topic, event_location, host_label, capacity, notes
    - active     TRUE/FALSE  (FALSE -> hidden; link shows "not active")
    - event_id   slug used in the URL & QR, e.g. nashville-fit-2026-04-29
    - region     city/area for grouping, e.g. Nashville
    - branch     host branch code, e.g. 101- Nashville
    - weekday    all-caps weekday (auto-filled here from the date; editable)
    - capacity   optional; notes optional (lunch/parking/etc.)

branches sheet: branch, territory_manager
    Kept for reference. Territory Manager is NOT auto-filled into registrations
    by default (the real data shows one branch -> several TMs). Flip
    FILL_TM_FROM_BRANCH in server.py if you ever want the lookup.
"""
from datetime import datetime
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font

REPO = Path(__file__).resolve().parent.parent  # tools/ -> repo root
OUT = REPO / "data" / "events.xlsx"

EVENT_COLUMNS = [
    "active", "event_id", "region", "branch", "weekday", "event_date",
    "start_time", "end_time", "topic", "event_location", "host_label",
    "capacity", "notes",
]

# weekday is filled automatically from event_date below (kept correct + editable).
# (active, event_id, region, branch, date, start, end, topic, location, host_label, capacity, notes)
EVENTS = [
    (True, "nashville-fit-2026-04-29", "Nashville", "101- Nashville", "2026-04-29",
     "09:00", "15:00", "FIT INSTALL & COMMISSIONING", "NASHVILLE BRANCH",
     "Hosted by M&A Supply — Nashville (Branch 101)", 24, "Lunch & parking provided."),
    (True, "columbia-fit-2026-06-17", "Columbia", "125- Columbia", "2026-06-17",
     "09:00", "15:00", "FIT INSTALL & COMMISSIONING", "COLUMBIA BRANCH",
     "Hosted by M&A Supply — Columbia (Branch 125)", 20, ""),
    (True, "murfreesboro-ductless-2026-07-15", "Murfreesboro", "107- Murfreesboro", "2026-07-15",
     "09:00", "12:00", "DUCTLESS SERVICE & TROUBLESHOOTING", "MURFREESBORO BRANCH",
     "Hosted by M&A Supply — Murfreesboro (Branch 107)", 16, "Bring your gauges."),
]

BRANCH_COLUMNS = ["branch", "territory_manager"]
BRANCHES = [
    ["101- Nashville", ""],
    ["107- Murfreesboro", ""],
    ["125- Columbia", ""],
    ["132- Cookeville", ""],
    ["145- Clarksville", ""],
    ["160- Knoxville", ""],
]


def _weekday(iso):
    return datetime.strptime(iso, "%Y-%m-%d").strftime("%A").upper()


def _event_rows():
    rows = []
    for (active, eid, region, branch, date, start, end, topic, loc, host, cap, notes) in EVENTS:
        rows.append([active, eid, region, branch, _weekday(date), date,
                     start, end, topic, loc, host, cap, notes])
    return rows


def write_sheet(ws, columns, rows):
    ws.append(columns)
    for c in ws[1]:
        c.font = Font(bold=True)
    for row in rows:
        ws.append(row)
    for i, col in enumerate(columns, start=1):
        vals = [str(col)] + [str(r[i - 1]) for r in rows]
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = min(max(len(v) for v in vals) + 2, 52)


def main():
    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws_events = wb.active
    ws_events.title = "events"
    write_sheet(ws_events, EVENT_COLUMNS, _event_rows())
    write_sheet(wb.create_sheet("branches"), BRANCH_COLUMNS, BRANCHES)
    wb.save(OUT)
    print(f"wrote {OUT}  ({len(EVENTS)} events, {len(BRANCHES)} branches)")


if __name__ == "__main__":
    main()
