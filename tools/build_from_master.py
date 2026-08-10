"""Compile the runtime data files from the ONE human-edited source:
`_data_in/MASTER COPY  Training Hub Info Phase 1.xlsx` (3 tabs).

    Tech Team Calendar  ->  data/OFFICIAL_CLASS_SCHEDULE.xlsx  (events + branches + closures)
    Tier 1 Dealers      ->  data/dealers.xlsx                  (John's "Results" export shape)
    M&A Personnel       ->  data/employees.xlsx                (Legal_Firstname, Legal_Lastname, Position, Work_Location)

Then run, in order, to finish the pipeline:
    ./.venv/bin/python tools/make_codes.py          # tm_keys.yaml + trainer_codes.yaml + access_codes.xlsx (real keys)
    ./.venv/bin/python tools/sync_chat_from_db.py   # reset chat history to REAL (zero until real signups)
    cd chat_bot && ../.venv/bin/python -m chat_rag.build_index

Runtime shapes are kept EXACTLY as the app expects; only the source of the files moves upstream to the MASTER.
Messy real-data handling: date typos, "City, ST"/"City ST" formats, "9:00 - 1:00 cst" business-hour times,
multi-trainer cells, and the internal San-Antonio conference (kept but marked inactive).

Run:  ./.venv/bin/python tools/build_from_master.py
"""
import csv
import re
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill

REPO = Path(__file__).resolve().parent.parent
def _newest_master():
    """The most recently modified MASTER workbook in _data_in/.
    John drops new versions with new names; we always build from the latest."""
    cands = sorted((REPO / "_data_in").glob("MASTER COPY*.xlsx"),
                   key=lambda p: p.stat().st_mtime, reverse=True)
    if not cands:
        raise SystemExit("No MASTER COPY*.xlsx found in _data_in/")
    return cands[0]


MASTER = _newest_master()
OUT_SCHED = REPO / "data" / "OFFICIAL_CLASS_SCHEDULE.xlsx"
OUT_DEALERS = REPO / "data" / "dealers.xlsx"
OUT_EMP = REPO / "data" / "employees.xlsx"
SIM = REPO / "chat_bot" / "data_global" / "classes_sim.csv"

DEFAULT_CAP = 20

# normalized topic -> (track/level, event_id short slug)
TOPIC_META = {
    "SANTE FE DEHUMIDIFIER TRAINING": ("Product", "dehum"),
    "RUUD COMMERCIAL PRODUCT":        ("Product", "ruud-commercial"),
    "RUUD 16/18 SIDE DISCHARGE":      ("Product", "ruud1618"),
    "HEAT PUMP DIAGNOSTICS":          ("Level 3", "heatpump"),
    "AIR FLOW & DUCT DESIGN":         ("Level 3", "airflow"),
    "ECONET ZONING":                  ("Product", "econet"),
    "FIT SERVICE":                    ("Level 1", "fitservice"),
    "DUCTLESS ONBOARDING DAIKIN":     ("Daikin", "ductless-onboarding"),
    "DUCTLESS + FIT TRAINING DAIKIN": ("Daikin", "ductless-fit"),
    "DAIKIN TECH SERVICES CONFERENCE": ("Internal", "daikin-conference"),
}

TRAINER_IDS = {
    "John Ward": "trainer_ward", "Shane McCloud": "trainer_mccloud",
    "Brian Thomson": "trainer_thomson", "Danny Barfield": "trainer_barfield",
    "Terry Jones": "trainer_jones",
}

BRANCH_CODES = {
    ("Nashville", "TN"): "101- Nashville", ("Murfreesboro", "TN"): "107- Murfreesboro",
    ("Columbia", "TN"): "125- Columbia", ("Cookeville", "TN"): "132- Cookeville",
    ("Knoxville", "TN"): "160- Knoxville", ("Chattanooga", "TN"): "150- Chattanooga",
    ("Huntsville", "AL"): "210- Huntsville", ("Birmingham", "AL"): "220- Birmingham",
    ("Memphis", "TN"): "310- Memphis", ("Little Rock", "AR"): "410- Little Rock",
}

# known source typos in the DATE column
DATE_FIX = {"10/2002026": "2026-10-20"}

EVENT_COLUMNS = [
    "active", "event_id", "region", "branch", "state", "weekday", "event_date",
    "start_time", "end_time", "timezone", "topic", "track", "trainer",
    "alt_trainer", "event_location", "host_label", "capacity", "notes",
]


# ---- parsing helpers --------------------------------------------------------
def norm_topic(s):
    return re.sub(r"\s+", " ", str(s or "").strip()).upper()


def slug(s):
    return re.sub(r"[^a-z0-9]+", "-", str(s).lower()).strip("-") or "tbd"


def parse_date(v):
    if hasattr(v, "date"):
        return v.date().isoformat()
    s = str(v or "").strip()
    if s in DATE_FIX:
        return DATE_FIX[s]
    for fmt in ("%Y-%m-%d", "%Y-%m-%d %H:%M:%S", "%m/%d/%Y", "%m/%d/%y"):
        try:
            return datetime.strptime(s, fmt).date().isoformat()
        except ValueError:
            pass
    return None


def parse_location(s):
    s = str(s or "").strip().rstrip(".")
    m = re.search(r"[, ]\s*([A-Za-z]{2})\.?$", s)
    if m:
        return s[:m.start()].strip().rstrip(",").strip(), m.group(1).upper()
    return s, ""


def _to24(t):
    m = re.match(r"(\d{1,2})(?::(\d{2}))?", t.strip())
    if not m:
        return ""
    h = int(m.group(1)); mins = m.group(2) or "00"
    if h != 12 and h <= 7:          # business hours: 1–7 read as PM, 8–12 as printed
        h += 12
    return f"{h:02d}:{mins}"


def parse_time(s):
    s = str(s or "").strip()
    if not s:
        return "", "", ""
    tz = ""
    m = re.search(r"\b(cst|cdt|est|edt|mst|pst)\b", s, re.I)
    if m:
        tz = m.group(1).upper(); s = (s[:m.start()] + s[m.end():]).strip()
    parts = re.split(r"\s*[-–]\s*", s)
    if len(parts) != 2:
        return "", "", tz
    return _to24(parts[0]), _to24(parts[1]), tz


# Texas is Central; a branch with no sibling row anywhere still needs a zone.
EXPLICIT_TZ = {"TX": "CST", "FL": "CST", "MS": "CST", "AL": "CST", "AR": "CST"}


def fill_timezones(rows):
    """Every timed class MUST state a timezone.

    The master only writes a zone when the trainer typed one, so most rows come
    back blank. A class shows the BRANCH's local time — never the viewer's — so
    a missing zone means a dealer can read the wrong hour and arrive late.

    Derive each blank from the same branch on other rows, then the same state.
    Returns (filled_count, still_blank) — the caller fails loudly on leftovers.
    """
    from collections import defaultdict
    by_branch, by_state = defaultdict(set), defaultdict(set)
    for r in rows:
        if r.get("timezone"):
            by_branch[r.get("branch", "")].add(r["timezone"])
            by_state[r.get("state", "")].add(r["timezone"])

    filled, blank = 0, []
    for r in rows:
        if r.get("timezone") or not r.get("start_time"):
            continue
        z = by_branch.get(r.get("branch", ""), set())
        if len(z) != 1:
            z = by_state.get(r.get("state", ""), set())
        tz = next(iter(z)) if len(z) == 1 else EXPLICIT_TZ.get(r.get("state", ""), "")
        if tz:
            r["timezone"] = tz
            filled += 1
        else:
            blank.append(r.get("event_id", "?"))
    return filled, blank


def weekday(iso):
    return datetime.strptime(iso, "%Y-%m-%d").strftime("%A").upper()


# ---- Tech Team Calendar -> class rows ---------------------------------------
def read_calendar():
    wb = load_workbook(MASTER, data_only=True)
    ws = wb["Tech Team Calendar"]
    rows, seen, skipped = [], set(), []
    for i, r in enumerate(ws.iter_rows(min_row=4, values_only=True), start=4):
        date_raw, loc_raw, desc_raw = r[1], r[2], r[3]
        if not (loc_raw and desc_raw):        # blank filler / spacer rows
            continue
        date = parse_date(date_raw)
        if not date:
            skipped.append((i, "bad date", str(date_raw))); continue
        city, state = parse_location(loc_raw)
        topic = norm_topic(desc_raw)
        track, short = TOPIC_META.get(topic, ("Class", slug(topic)[:14]))
        start, end, tz = parse_time(r[4])
        trainer = re.sub(r"\s+", " ", str(r[5] or "").strip()) or "TBD"
        alt = str(r[6] or "").strip()
        active = track != "Internal"
        eid = f"{slug(city)}-{short}-{date}"
        if eid in seen:
            eid = f"{eid}-{slug(state) or 'x'}"
        seen.add(eid)
        branch = BRANCH_CODES.get((city, state), f"{city}, {state}".strip().strip(","))
        flags = []
        if not (start and end):
            flags.append("Time TBD")
        if trainer.upper() == "TBD":
            flags.append("Trainer TBD")
        note = []
        if alt:
            note.append(f"Alt/Rep: {alt}")
        if not active:
            note.append("Internal team event — not dealer-facing")
        rows.append({
            "active": active, "event_id": eid, "region": city, "branch": branch,
            "state": state, "weekday": weekday(date), "event_date": date,
            "start_time": start, "end_time": end, "timezone": tz,
            "topic": topic, "track": track, "trainer": trainer, "alt_trainer": alt,
            "event_location": f"{city.upper()} BRANCH" if city else "TBD",
            "host_label": f"Hosted by M&A Supply — {city}, {state}".rstrip(", "),
            "capacity": DEFAULT_CAP, "notes": "; ".join(note), "status": " / ".join(flags),
        })
    wb.close()
    return rows, skipped


def write_schedule(rows):
    wb = Workbook()
    ws = wb.active; ws.title = "events"
    head = PatternFill("solid", fgColor="0A2540")
    ws.append(EVENT_COLUMNS + ["status"])
    for c in ws[1]:
        c.font = Font(bold=True, color="FFFFFF"); c.fill = head
    tbd = PatternFill("solid", fgColor="FFF3CD"); off = PatternFill("solid", fgColor="EEEEEE")
    for r in rows:
        ws.append([r[k] for k in EVENT_COLUMNS] + [r["status"]])
        row = ws[ws.max_row]
        if not r["active"]:
            for c in row:
                c.fill = off
        elif r["status"]:
            row[-1].fill = tbd
    ws.freeze_panes = "A2"
    ws2 = wb.create_sheet("branches")
    ws2.append(["branch", "territory_manager"])
    ws2["A1"].font = ws2["B1"].font = Font(bold=True)
    for b in sorted({r["branch"] for r in rows}):
        ws2.append([b, ""])
    for s in wb.worksheets:
        for col in s.columns:
            w = max((len(str(c.value or "")) for c in col), default=8)
            s.column_dimensions[col[0].column_letter].width = min(w + 2, 60)
    wb.save(OUT_SCHED)


def write_sim(rows):
    with SIM.open("w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["event_id", "region", "branch", "tm_id", "topic", "level",
                    "event_date", "start_time", "end_time", "capacity", "seats_taken",
                    "seats_remaining", "location", "trainer_id", "trainer_name"])
        for r in rows:
            tid = TRAINER_IDS.get(r["trainer"], "")
            w.writerow([r["event_id"], r["region"], r["branch"], "", r["topic"], r["track"],
                        r["event_date"], r["start_time"], r["end_time"], r["capacity"],
                        0, r["capacity"], r["event_location"], tid, r["trainer"]])


# ---- Tier 1 Dealers -> dealers.xlsx (John's Results shape) -------------------
DEALER_SHEETS = ("Top Tier Dealers", "Tier 1 Dealers")   # renamed Aug 2026


def build_dealers():
    wb = load_workbook(MASTER, data_only=True)
    name = next((s for s in DEALER_SHEETS if s in wb.sheetnames), None)
    if not name:
        raise SystemExit(f"No dealer sheet found. Looked for {DEALER_SHEETS}, "
                         f"workbook has {wb.sheetnames}")
    ws = wb[name]
    dealers, seen = [], set()
    # Shape-tolerant: banner/header rows have moved between versions, so keep any
    # row whose first cell is a numeric Customer ID and second cell is a name.
    for r in ws.iter_rows(min_row=1, values_only=True):
        if not r or len(r) < 2:
            continue
        cid, nm = r[0], r[1]
        if cid is None or not str(nm or "").strip():
            continue
        cid = str(cid).strip()
        if not cid.isdigit():             # skips dates, banners, header text
            continue
        if cid in seen:
            continue
        seen.add(cid)
        rep = str(r[2] or "").strip() if len(r) > 2 else ""
        branch = str(r[3] or "").strip() if len(r) > 3 else ""
        dealers.append((cid, str(nm).strip(), rep, branch))
    wb.close()
    out = Workbook(); s = out.active; s.title = "Results"
    # Clean table: header on ROW 1 (no empty banner rows). Customer ID is the primary
    # key — it becomes each registration's Account Number (src/registrations.py). The
    # 4th "Branch" column is ours (not John's export) — kept for the personnel
    # Access Hub dealer region; src/dealers.py only reads cols A-C, so it's harmless there.
    from openpyxl.styles import Font
    s.append(["Customer ID", "Customer Name", "Sales Rep Name (Cust)", "Branch"])   # header dealers.py detects
    for cid, name, rep, branch in dealers:
        s.append([cid, name, rep, branch])                   # rep stays "Last, First" (dealers.py flips it)
    for cell in s[1]:                                        # bold the header row
        cell.font = Font(bold=True)
    s.freeze_panes = "A2"                                    # keep header visible while scrolling
    for col, width in zip("ABCD", (14, 40, 26, 16)):        # readable column widths
        s.column_dimensions[col].width = width
    out.save(OUT_DEALERS)
    return len(dealers)


# ---- M&A Personnel -> employees.xlsx (order make_codes.py expects) ----------
def build_employees():
    wb = load_workbook(MASTER, data_only=True)
    ws = wb["M&A Personnel"]
    hdr = [str(c.value or "").strip() for c in ws[1]]
    idx = {h: i for i, h in enumerate(hdr)}
    li, fi, pi, wi = (idx["Legal_Lastname"], idx["Legal_Firstname"],
                      idx["Position"], idx["Work_Location"])
    people = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        if not (r[li] or r[fi]):
            continue
        people.append((str(r[fi] or "").strip(), str(r[li] or "").strip(),
                       str(r[pi] or "").strip(), str(r[wi] or "").strip()))
    wb.close()
    out = Workbook(); s = out.active; s.title = "Employees"
    s.append(["Legal_Firstname", "Legal_Lastname", "Position", "Work_Location"])
    for row in people:
        s.append(list(row))
    out.save(OUT_EMP)
    return len(people)


def main():
    rows, skipped = read_calendar()
    n_tz, tz_blank = fill_timezones(rows)
    if n_tz:
        print(f"  timezone: derived {n_tz} missing value(s) from branch/state")
    if tz_blank:
        raise SystemExit(f"ABORT: {len(tz_blank)} class(es) have no timezone and none "
                         f"could be derived: {tz_blank[:6]}")
    write_schedule(rows)
    write_sim(rows)
    n_dealers = build_dealers()
    n_emp = build_employees()

    active = [r for r in rows if r["active"]]
    from collections import Counter
    months = Counter(r["event_date"][:7] for r in active)
    print(f"{OUT_SCHED.name}: {len(rows)} classes ({len(active)} active, {len(rows)-len(active)} internal)")
    print("  by month:", dict(sorted(months.items())))
    print(f"  topics: {sorted({r['topic'] for r in active})}")
    print(f"{OUT_DEALERS.name}: {n_dealers} dealers")
    print(f"{OUT_EMP.name}: {n_emp} personnel")
    if skipped:
        print("  ⚠ skipped rows:", skipped)
    print("\nNext: make_codes.py  ->  sync_chat_from_db.py  ->  chat_rag.build_index")


if __name__ == "__main__":
    main()
