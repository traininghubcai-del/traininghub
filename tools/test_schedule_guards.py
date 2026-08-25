"""Proof that the catalog write path refuses a bad date or a backwards clock.

Runs against a THROWAWAY copy of the catalog in a temp directory — never the
live workbook. The copy is made first, the env is pointed at it, and only then
is anything under src/ imported, because config resolves EVENTS_XLSX at import
time.

    python3 tools/test_schedule_guards.py
"""
import os
import shutil
import sys
import tempfile
from datetime import date, timedelta
from pathlib import Path

HERE = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(HERE))

# ---- sandbox the data BEFORE config is imported ---------------------------
LIVE = Path(os.environ.get("EVENTS_XLSX") or HERE / "data" / "OFFICIAL_CLASS_SCHEDULE.xlsx")
SANDBOX = Path(tempfile.mkdtemp(prefix="schedule-guards-"))
COPY = SANDBOX / "OFFICIAL_CLASS_SCHEDULE.xlsx"
shutil.copy2(LIVE, COPY)
os.environ["APP_DATA_DIR"] = str(SANDBOX)
os.environ["EVENTS_XLSX"] = str(COPY)

from src.manage import create_class, update_class  # noqa: E402
from src.hub_modes import create_class as hub_create, save_class as hub_save  # noqa: E402
from src.hub_modes import DEFAULT_CODE  # noqa: E402

TODAY = date.today()
YESTERDAY = (TODAY - timedelta(days=1)).isoformat()
TOMORROW = (TODAY + timedelta(days=1)).isoformat()


class FakeRepo:
    """Only what the write path touches — no DB, no registrations."""
    def seats_taken(self, event_id):
        return 0


REPO = FakeRepo()
PASSED, FAILED = [], []


def fingerprint():
    """(size, mtime_ns, every event_date in the sheet) — any write moves this."""
    from openpyxl import load_workbook
    st = COPY.stat()
    wb = load_workbook(COPY, read_only=True, data_only=True)
    ws = wb["events"]
    headers = [c.value for c in ws[1]]
    di, ii = headers.index("event_date"), headers.index("event_id")
    rows = tuple((r[ii], str(r[di])) for r in ws.iter_rows(min_row=2, values_only=True) if r[ii])
    wb.close()
    return st.st_size, st.st_mtime_ns, rows


def check(name, condition, detail=""):
    (PASSED if condition else FAILED).append(name)
    print(f"  {'PASS' if condition else 'FAIL'}  {name}" + (f"   [{detail}]" if detail else ""))


def a_past_class():
    """An event_id already dated before today — needed for the 'edit a finished
    class' cases. There is one in the real catalog; if not, make one by writing
    the row directly (bypassing the guard we're testing, on purpose)."""
    from openpyxl import load_workbook
    wb = load_workbook(COPY)
    ws = wb["events"]
    headers = [c.value for c in ws[1]]
    di, ii = headers.index("event_date") + 1, headers.index("event_id") + 1
    today_s = TODAY.isoformat()
    for r in range(2, ws.max_row + 1):
        eid = str(ws.cell(row=r, column=ii).value or "").strip()
        if eid and str(ws.cell(row=r, column=di).value or "")[:10] < today_s:
            wb.close()
            return eid
    # none in the sheet — backdate the first row at the file level
    ws.cell(row=2, column=di).value = (TODAY - timedelta(days=30)).isoformat()
    wb.save(COPY)
    eid = str(ws.cell(row=2, column=ii).value or "").strip()
    wb.close()
    return eid


def a_future_class():
    from src.manage import list_classes
    rows = [c for c in list_classes(REPO)["classes"] if not c["is_past"]]
    return rows[0]["event_id"] if rows else None


def main():
    print(f"catalog copy : {COPY}")
    print(f"today        : {TODAY}   yesterday: {YESTERDAY}   tomorrow: {TOMORROW}\n")

    past_id = a_past_class()
    future_id = a_future_class()
    print(f"past class   : {past_id}")
    print(f"future class : {future_id}\n")

    # 1 -- edit a date into the past, and prove the file did not move
    before = fingerprint()
    r = update_class(REPO, future_id, {"event_date": YESTERDAY})
    check("1. update_class -> yesterday is refused",
          r.get("ok") is False and "already passed" in r.get("error", ""), r.get("error", ""))
    check("1b. workbook untouched by the refused edit", fingerprint() == before)

    # 2 -- edit a date to tomorrow
    r = update_class(REPO, future_id, {"event_date": TOMORROW})
    check("2. update_class -> tomorrow is accepted",
          r.get("ok") is True and r["class"]["event_date"] == TOMORROW,
          r.get("error") or r["class"]["event_date"])

    # 3 -- create dated yesterday
    r = create_class(REPO, {"region": "Testville", "topic": "GUARD TEST A",
                            "event_date": YESTERDAY, "branch": "99 - Testville, AR"})
    check("3. create_class -> yesterday is refused",
          r.get("ok") is False and "already passed" in r.get("error", ""), r.get("error", ""))

    # 4 -- create dated tomorrow
    r = create_class(REPO, {"region": "Testville", "topic": "GUARD TEST B",
                            "event_date": TOMORROW, "branch": "99 - Testville, AR"})
    check("4. create_class -> tomorrow is accepted",
          r.get("ok") is True, r.get("error", ""))
    made = (r.get("class") or {}).get("event_id", "")

    # 5 -- end before start
    r = update_class(REPO, future_id, {"start_time": "14:00", "end_time": "09:00"})
    check("5. start 14:00 / end 09:00 is refused",
          r.get("ok") is False and r.get("error") == "End time must be after start time.",
          r.get("error", ""))

    # 5b -- change ONLY the end time, into a value before the stored start
    update_class(REPO, future_id, {"start_time": "10:00", "end_time": "15:00"})
    r = update_class(REPO, future_id, {"end_time": "08:00"})
    check("5b. end-only edit is checked against the STORED start",
          r.get("ok") is False and r.get("error") == "End time must be after start time.",
          r.get("error", ""))

    # 6 -- a forwards pair
    r = update_class(REPO, future_id, {"start_time": "09:00", "end_time": "14:00"})
    check("6. start 09:00 / end 14:00 is accepted", r.get("ok") is True, r.get("error", ""))

    # 7 -- edit a field on an ALREADY past class, with no date in the patch
    r = update_class(REPO, past_id, {"trainer": "GUARD TEST TRAINER"})
    check("7. trainer-only edit on a finished class still saves",
          r.get("ok") is True and r["class"]["trainer"] == "GUARD TEST TRAINER",
          r.get("error", ""))
    # ...and moving that finished class FORWARD is allowed (fixing a typo)
    r = update_class(REPO, past_id, {"event_date": TOMORROW})
    check("7b. a finished class may be moved to a future date",
          r.get("ok") is True, r.get("error", ""))
    # ...but not to another past day
    r = update_class(REPO, past_id, {"event_date": YESTERDAY})
    check("7c. ...and not to another past date",
          r.get("ok") is False and "already passed" in r.get("error", ""), r.get("error", ""))

    # 8 -- the API layer, exactly as routes.py calls it (a forged POST body)
    before = fingerprint()
    r = hub_save(REPO, future_id, "admin", DEFAULT_CODE, {"event_date": YESTERDAY})
    check("8. /api/hub/save-class forged with a past date is refused",
          r.get("ok") is False and "already passed" in r.get("error", ""), r.get("error", ""))
    check("8b. workbook untouched by the forged save", fingerprint() == before)
    r = hub_create(REPO, "admin", DEFAULT_CODE,
                   {"branch": "99 - Testville, AR", "topic": "GUARD TEST C",
                    "event_date": YESTERDAY, "timezone": "CST"})
    check("8c. /api/hub/create-class forged with a past date is refused",
          r.get("ok") is False and "already passed" in r.get("error", ""), r.get("error", ""))

    # 9 -- garbage dates and garbage clocks
    r = update_class(REPO, future_id, {"event_date": "09/01/2026"})
    check("9. a non-ISO date is refused",
          r.get("ok") is False and r.get("error") == "Date must be YYYY-MM-DD.", r.get("error", ""))
    r = update_class(REPO, future_id, {"start_time": "quarter past nine"})
    check("9b. an unreadable time is refused",
          r.get("ok") is False and "must be a time like" in r.get("error", ""), r.get("error", ""))

    # 10 -- blank times stay legal (TBD classes are a real thing in this catalog)
    r = update_class(REPO, future_id, {"start_time": "", "end_time": ""})
    check("10. blank start/end (TBD) is still allowed", r.get("ok") is True, r.get("error", ""))

    print(f"\n{len(PASSED)} passed, {len(FAILED)} failed")
    if made:
        print(f"(created in the sandbox copy only: {made})")
    print(f"sandbox left at {SANDBOX} — delete it freely")
    return 1 if FAILED else 0


if __name__ == "__main__":
    sys.exit(main())
