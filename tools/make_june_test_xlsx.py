"""Generate data/events_june_test.xlsx — a TEMP catalog of 10 June test classes.

Same structure as data/events.xlsx (events + branches sheets). This is throwaway
test data and is kept separate from the real catalog. Serve it with:

    EVENTS_XLSX=data/events_june_test.xlsx python3 server.py
    EVENTS_XLSX=data/events_june_test.xlsx python3 make_qr.py

weekday is derived from each date so it can't be mistyped.
Reset:  python3 make_june_test_xlsx.py
"""
from datetime import datetime
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font

REPO = Path(__file__).resolve().parent.parent  # tools/ -> repo root
OUT = REPO / "data" / "events_june_test.xlsx"

EVENT_COLUMNS = [
    "active", "event_id", "region", "branch", "weekday", "event_date",
    "start_time", "end_time", "topic", "event_location", "host_label",
    "capacity", "notes",
]

# (active, event_id, region, branch, date, start, end, topic, location, host_label, capacity)
JUNE = [
    (True, "nashville-fit-2026-06-03", "Nashville", "101- Nashville", "2026-06-03", "09:00", "15:00",
     "FIT INSTALL & COMMISSIONING", "NASHVILLE BRANCH", "Hosted by M&A Supply — Nashville (Branch 101)", 24),
    (True, "nashville-mini-split-2026-06-10", "Nashville", "101- Nashville", "2026-06-10", "09:00", "14:00",
     "MINI SPLIT INSTALL", "NASHVILLE BRANCH", "Hosted by M&A Supply — Nashville (Branch 101)", 24),
    (True, "murfreesboro-airflow-2026-06-05", "Murfreesboro", "107- Murfreesboro", "2026-06-05", "09:00", "14:00",
     "AIR FLOW & DUCT DESIGN", "MURFREESBORO BRANCH", "Hosted by M&A Supply — Murfreesboro (Branch 107)", 20),
    (True, "murfreesboro-heatpump-2026-06-19", "Murfreesboro", "107- Murfreesboro", "2026-06-19", "09:00", "14:00",
     "HEAT PUMP DIAGNOSTICS", "MURFREESBORO BRANCH", "Hosted by M&A Supply — Murfreesboro (Branch 107)", 20),
    (True, "columbia-mini-split-2026-06-04", "Columbia", "125- Columbia", "2026-06-04", "09:00", "14:00",
     "MINI SPLIT INSTALL", "COLUMBIA BRANCH", "Hosted by M&A Supply — Columbia (Branch 125)", 18),
    (True, "columbia-airflow-2026-06-18", "Columbia", "125- Columbia", "2026-06-18", "09:00", "14:00",
     "AIR FLOW & DUCT DESIGN", "COLUMBIA BRANCH", "Hosted by M&A Supply — Columbia (Branch 125)", 18),
    (True, "cookeville-heatpump-2026-06-09", "Cookeville", "132- Cookeville", "2026-06-09", "09:00", "14:00",
     "HEAT PUMP DIAGNOSTICS", "COOKEVILLE BRANCH", "Hosted by M&A Supply — Cookeville (Branch 132)", 16),
    # NOTE: original seed row was missing its branch; set to 132- Cookeville.
    (True, "cookeville-fit-2026-06-23", "Cookeville", "132- Cookeville", "2026-06-23", "09:00", "15:00",
     "FIT INSTALL & COMMISSIONING", "COOKEVILLE BRANCH", "Hosted by M&A Supply — Cookeville (Branch 132)", 16),
    (True, "knoxville-mini-split-2026-06-11", "Knoxville", "160- Knoxville", "2026-06-11", "09:00", "14:00",
     "MINI SPLIT INSTALL", "KNOXVILLE BRANCH", "Hosted by M&A Supply — Knoxville (Branch 160)", 20),
    (True, "knoxville-airflow-2026-06-25", "Knoxville", "160- Knoxville", "2026-06-25", "09:00", "14:00",
     "AIR FLOW & DUCT DESIGN", "KNOXVILLE BRANCH", "Hosted by M&A Supply — Knoxville (Branch 160)", 20),
]

BRANCH_COLUMNS = ["branch", "territory_manager"]
BRANCHES = [
    ["101- Nashville", ""], ["107- Murfreesboro", ""], ["125- Columbia", ""],
    ["132- Cookeville", ""], ["145- Clarksville", ""], ["160- Knoxville", ""],
]


def _weekday(iso):
    return datetime.strptime(iso, "%Y-%m-%d").strftime("%A").upper()


def _event_rows():
    rows = []
    for (active, eid, region, branch, date, start, end, topic, loc, host, cap) in JUNE:
        rows.append([active, eid, region, branch, _weekday(date), date,
                     start, end, topic, loc, host, cap, ""])
    return rows


def write_sheet(ws, columns, rows):
    ws.append(columns)
    for c in ws[1]:
        c.font = Font(bold=True)
    for row in rows:
        ws.append(row)
    for i, col in enumerate(columns, start=1):
        vals = [str(col)] + [str(r[i - 1]) for r in rows]
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = min(max(len(v) for v in vals) + 2, 52)


def main():
    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "events"
    write_sheet(ws, EVENT_COLUMNS, _event_rows())
    write_sheet(wb.create_sheet("branches"), BRANCH_COLUMNS, BRANCHES)
    wb.save(OUT)
    print(f"wrote {OUT}  ({len(JUNE)} events, {len(BRANCHES)} branches)")


if __name__ == "__main__":
    main()
